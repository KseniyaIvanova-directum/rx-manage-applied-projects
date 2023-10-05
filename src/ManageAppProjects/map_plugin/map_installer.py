# coding: utf-8
""" Модуль плагина для управления прикладными проектами  """
import pathlib
from pprint import pprint, pformat
from typing import Optional, Dict, Any, List, OrderedDict
import termcolor
import time
import shutil
from pathlib import PurePath, Path
import os
import sys
import json
from ui_installer.lib.click import pause
from ruamel.yaml import CommentedMap, CommentedSeq
from openpyxl import *

from fire.formatting import Bold

from components.base_component import BaseComponent
from components.component_manager import component
from py_common.logger import log
from common_plugin import yaml_tools
from sungero_deploy.all import All
from sungero_deploy.scripts_config import get_config_model
from sungero_deploy.tools.sungerodb import SungeroDB
from py_common import io_tools, process
from sungero_deploy.scripts_config import Config
from common_plugin import git_tools


MANAGE_APPLIED_PROJECTS_ALIAS = 'map'
DELETE_MARKER = "_delete_%$#@"
ESS_SOURCE = "ЛК"
APP_SOURCE = "ПЧ"
SETTINGS_SOURCE = "Настройки"

#region service function

def _get_rx_version(need_short: bool = True) -> str:
    """Вернуть версию RX
    """
    # версия 4.2. Информация о билде прикладной хранится в version.txt
    version_dict = yaml_tools.load_yaml_from_file(_get_check_file_path("etc\\_builds\\version.txt"))
    applied_builds_version = version_dict["builds"].get("applied_builds", None)
    if applied_builds_version is not None:
        return applied_builds_version["version"]

    with open(_get_check_file_path("etc\\_builds\\DirectumRX\\manifest.json"),  'r', encoding='utf-8') as manifest_json:
        data = " ".join(manifest_json.readlines())
        manifest_dict = json.loads(data)
        return manifest_dict["version"]

def _copy_database_mssql(config: Config, src_db_name: str, dst_db_name: str) -> None:
    """Создать копию базы данных на Microsoft SQL Server.

    Args:
        config: конфиг Sungero.
        src_db_name: исходная БД.
        dst_db_name: целевая БД.
    """
    log.info(f'Create database backup: "{src_db_name}".')

    # найти каталог для резервных копий
    # в DirectumLauncher 4.4 изменилось имя функции, поэтому пробуем оба варианта
    database_folder = None
    try:
        # версия 4.4
        from sungero_tenants.dbtools import get_mssql_database_folder
        database_folder = get_mssql_database_folder(config, src_db_name)
    except:
        pass
    if database_folder is None:
        try:
            # до версии 4.4
            from sungero_tenants.dbtools import get_database_folder
            database_folder = get_database_folder(config, src_db_name)
        except:
            pass
    if database_folder is None:
        try:
            # версия 4.5
            from platform_plugin.sungero_tenants.dbtools import get_mssql_database_folder
            database_folder = get_mssql_database_folder(config, src_db_name)
        except:
            pass
    if database_folder is None:
        raise ValueError(f"Не удалось найти функцию для получения имени каталога резервных копий.")

    command_text = f"""
        -- ============ копипаста из dbtools.create_database_backup() ============
        declare @DatabaseName sysname = '{src_db_name}'
        declare @DatabaseFolder nvarchar(255) = '{database_folder}'

        -- Получить путь к последнему полному бэкапу базы.
        declare @FullBackupPath nvarchar(255) = ''
        set @FullBackupPath = @DatabaseFolder + @DatabaseName + '_' + replace(cast(newid() as varchar(36)), '-', '') + '.full'

        declare @BackupName nvarchar(255) = 'Backup created by tenant manage script'
        backup database @DatabaseName to disk = @FullBackupPath with copy_only, init, name = @BackupName
        if @@ERROR <> 0
        begin
          print(@FullBackupPath)
          return
        end
        print('!Создана полная резервная копия "' + @FullBackupPath + '"')
        select @FullBackupPath

        -- =========== копипаста из dbtools.create_database_from_backup() ============
        declare @NewDatabaseName sysname = '{dst_db_name}'

        print('!Создание из резервной копии: "' + @FullBackupPath + '"')
        -- Сформировать список файлов эталонной базы данных для запроса восстановления из бэкапа.
        declare @productver VARCHAR(50) = (SELECT CAST(SERVERPROPERTY('productversion') AS VARCHAR(50)))
        declare @version int = CAST(LEFT(@productver, CHARINDEX('.', @productver)-1) AS INT)

        create table #BackupFiles (LogicalName nvarchar(128), PhysicalName nvarchar(260), Type char(1), FileGroupName nvarchar(120),
            Size numeric(20, 0), MaxSize numeric(20, 0), FileID bigint, CreateLSN numeric(25,0), DropLSN numeric(25,0),
            UniqueID uniqueidentifier, ReadOnlyLSN numeric(25,0), ReadWriteLSN numeric(25,0), BackupSizeInBytes bigint,
            SourceBlockSize int, FileGroupID int, LogGroupGUID uniqueidentifier, DifferentialBaseLSN numeric(25,0),
            DifferentialBaseGUID uniqueidentifier, IsReadOnly bit, IsPresent bit, TDEThumbprint varbinary(32));
        -- С версии SQL2016 появилась новая колонка.
        if @version > 12
        begin
        alter table #BackupFiles
            add SnapshotUrl nvarchar(2083) NULL;
        end
        insert into #BackupFiles
        exec('restore filelistonly from disk = ''' + @FullBackupPath + '''')
        if @@ERROR <> 0
        return
        declare @MoveStatement nvarchar(4000) = ''
        select
        @MoveStatement = @MoveStatement + ', move ''' + LogicalName + ''' to ''' +
            @DatabaseFolder + @NewDatabaseName +
            case
            when Type = 'D' then '.mdf'
            when Type = 'L' then '_log.ldf'
            when Type = 'F' then '\FullTextData'
            end + ''''
        from
        #BackupFiles
        drop table #BackupFiles
        print(@MoveStatement)
        -- Восстановить новую базу из бэкапа эталонной.
        exec('restore database [' + @NewDatabaseName + '] from disk = ''' + @FullBackupPath + ''' with recovery, replace ' + @MoveStatement)
        if @@ERROR = 0
            print('!База данных "' + @NewDatabaseName + '" создана')

        -- Удалить созданную резервную копию
        declare @Command varchar(4000) = ''
        set @Command = 'del "' + @FullBackupPath + '"'
        exec master..xp_cmdshell @Command
        if @@ERROR = 0
            print('!Файл созданной резервной копии удален')
        """
    if 'platform_plugin.sungero_tenants.dbtools' in sys.modules:
        from platform_plugin.sungero_tenants.dbtools import ENABLE_XP_CMDSHELL # 4.5
    else:
        from sungero_tenants.dbtools import ENABLE_XP_CMDSHELL # 4.2-4.4

    result = SungeroDB(config).execute_command(ENABLE_XP_CMDSHELL.format(command_text), return_results=True)
    log.info(f'Database copied: {result}')

def _copy_database_postgresql(src_sungero_config: Any, src_db_name: str, dst_db_name: str):
    """Создать копию базы данных на PostgreSQL.

    Args:
        config: конфиг Sungero в виде yaml.
        src_db_name: исходная БД.
        dst_db_name: целевая БД.
    """
    postgree_path = _get_map_settings(config=src_sungero_config, param_name="postgresql_bin", is_required=True)
    # достать параметры подключения к Postgree
    connection_string_yml = src_sungero_config["common_config"]["CONNECTION_STRING"].split(";")
    server = ""
    port = ""
    username = ""
    for param in connection_string_yml:
        p = param.split("=")
        if p[0].lower() == "server":
            server = p[1]
        if p[0].lower() == "user id":
            username = p[1]
        if p[0].lower() == "port":
            port = p[1]
    # сформировать строку подключения к серверу в зависимости от используемого типа аутентификации
    connection_string = f'--host={server} --port={port} --username={username} --no-password'
    cmd = f'"{postgree_path}\\createdb.exe" {connection_string} {dst_db_name}'
    log.debug(f'{cmd}')
    exit_code = process.try_execute(cmd, encoding='cp1251') #cp1251  utf-8
    if exit_code != 0:
        raise IOError(f'Ошибка при создании БД')
    cmd = f'"{postgree_path}\\pg_dump.exe" {connection_string} {src_db_name} | "{postgree_path}\\psql.exe" -q {connection_string} {dst_db_name}'
    exit_code = process.try_execute(cmd, encoding='cp1251')
    if exit_code != 0:
        raise IOError(f'Ошибка при копировании данных БД')

def _colorize(x, color, attrs):
    return termcolor.colored(x, color=color, attrs=attrs)
def _colorize_green(x):
    return _colorize(x, color="green", attrs=["bold"])
def _colorize_red(x):
    return _colorize(x, color="red", attrs=["bold"])
def _colorize_cyan(x):
    return _colorize(x, color="cyan", attrs=["bold"])

def _get_url(config) -> None:
    """Вернуть  url для открытия веб-клиента текущего инстанса"""
    vars = config.variables
    srv_cfgs = config.services_config
    return f'{vars["protocol"]}://{vars["host_fqdn"]}:{vars["http_port"]}/{srv_cfgs["SungeroWebServer"]["WEB_HOST_PATH_BASE"]}/#'

def _show_config(config_path):
    config = yaml_tools.load_yaml_from_file(_get_check_file_path(config_path))
    vars = config.get("variables")
    repos = config.get("services_config").get("DevelopmentStudio").get('REPOSITORIES').get("repository")
    maxlen = 0
    for repo in repos:
        if maxlen < len(repo.get("@folderName")):
            maxlen = len(repo.get("@folderName"))
    log.info(Bold(f'Назначение:          {vars.get("purpose")}'))
    if vars.get("project_config_path") is not None:
        log.info(f'project_config_path: {_colorize_green(vars.get("project_config_path"))}')
    log.info(f'database:            {_colorize_green(vars.get("database"))}')
    log.info(f'home_path:           {_colorize_green(vars.get("home_path"))}')
    log.info(f'home_path_src:       {_colorize_green(vars.get("home_path_src"))}')
    log.info('repositories:')
    repos_str = []
    maxlen_folder = 0
    maxlen_status = 0
    for repo in repos:
        folder_str = f'folder: {_colorize_green(repo.get("@folderName")):}'
        solutiontype_str = f'solutiontype: {_colorize_green(repo.get("@solutionType"))}'
        url_str = f'url: {_colorize_green(repo.get("@url"))}'
        status_str = f'status: {repo_info(vars.get("home_path_src"), repo.get("@folderName"))}'
        repos_str.append({"folder": folder_str,
                          "solutiontype": solutiontype_str,
                          "url": url_str,
                          "status": status_str})
        maxlen_folder = len(folder_str) if maxlen_folder < len(folder_str) else maxlen_folder
        maxlen_status = len(status_str) if maxlen_status < len(status_str) else maxlen_status

    for repo_str in repos_str:
        log.info(f'  {repo_str["folder"].ljust(maxlen_folder)} {repo_str["status"].ljust(maxlen_status)} {repo_str["solutiontype"]} {repo_str["url"]}')

def _get_check_file_path(config_path: str) -> Path:
    if not config_path:
        raise ValueError("config_path does not set.")
    p_config_path = Path(config_path)
    if not p_config_path.is_file():
        log.error(f'Файл {config_path} не найден.')
        raise FileNotFoundError(f"'config_path' file not found: '{config_path}'")
    return p_config_path

def _get_full_path(root: str, relative: str) -> str:
    """Вычислить полный путь. Если параметр relative содержит абаслютный путь - то возвращает значение этого параметра.
    В противном случае возвращается root+relative.
    """
    if Path(relative).is_absolute():
        return str(relative)
    else:
        return str(PurePath(root, relative))

def _generate_empty_config_by_template(new_config_path: str, template_config: str) -> None:
    """ Создать новый файл конфига по шаблону """
    p_config_path = pathlib.Path(new_config_path)
    if not p_config_path.exists():
        with open(new_config_path, 'w', encoding='utf-8') as f:
            f.write(template_config)
        log.info(_colorize_green(f'Создан файл {new_config_path}.'))
    else:
        log.error(f'Файл {new_config_path} уже существует.')

def _update_sungero_config(project_config_path, sungero_config_path):
    """Преобразовать текущий config.yml в соотвтетствии с указанным конфигом проекта.
    Преобразование выполняется без сохранения на диске

    Args:
        * project_config_path - путь к конфигу проекта
        * sungero_config_path - путь к config.yml

    Return:
        * преоразованный конфиг
    """
    src_config = yaml_tools.load_yaml_from_file(project_config_path)
    dst_config = yaml_tools.load_yaml_from_file(sungero_config_path)
    dst_config["services_config"]["DevelopmentStudio"]['REPOSITORIES']["repository"]  = src_config["services_config"]["DevelopmentStudio"]['REPOSITORIES']["repository"].copy()
    dst_config["variables"]["purpose"] = src_config["variables"]["purpose"]
    dst_config["variables"]["database"] = src_config["variables"]["database"]
    dst_config["variables"]["home_path"] = src_config["variables"]["home_path"]
    dst_config["variables"]["home_path_src"]  = src_config["variables"]["home_path_src"]
    # костыль по быстрому, чтобы project_config_path была нужного типа
    dst_config["variables"]["project_config_path"]  = dst_config["variables"]["database"]
    dst_config["variables"]["project_config_path"] = project_config_path

    return dst_config

def _get_map_settings(config_path: str = None, config: Any = None, param_name: str = None, is_required: bool = False, default_value: Any = None) -> Any:
    """Получить значение параметра компоненты Manage Applied Projects из config.yml

    Args:
        config_path: str - путь к конфигу
        config: str - сам конфиг. config_path и config - взаимоисключающие параметры, config имеет приоритет
        param_name: str - имя параметра
        is_required: bool = False - если True, то при отсутствии параметра в config.yml будет выброшено исключение
        default_value: Any = None - значение по умолчанию. Если is_required=False и параметра нет в config.yml, то вернется default_value
    """
    if config is None:
        if config_path is not None:
            config = yaml_tools.load_yaml_from_file(_get_check_file_path(config_path)) #get_config_model(config_path)
        else:
            raise AssertionError('Должен быть либо указан параметр config, либо config_path')
    if "manage_applied_projects" in config:
        manage_applied_projects_config = config.get("manage_applied_projects", None)
        if param_name in manage_applied_projects_config:
            return manage_applied_projects_config.get(param_name)
        else:
            if is_required:
                raise AssertionError(f'В config.yml отсутствует параметр manage_applied_projects -> {param_name}')
            else:
                return default_value
    else:
        if is_required:
            raise AssertionError('В config.yml отсутствует раздел "manage_applied_projects"')
        else:
            return default_value

def _run_dds(config_path: str, need_run: bool, confirm: bool) -> None:
    """Запустить DDS, если попросили об этом (параметр need_run) или в конфиге настроен запуск по умолчанию.
    Если явно попросили запустить DDS, а он не установлен, то DDS не будет запущен, а в log выведется сообщение.
    Если параметр confirm установлен в True, то перед запуском DDS будет выведен запрос на подтверждение запуска.
    """
    if need_run or (need_run is None and _get_map_settings(config_path=config_path,
                                                            param_name="run_dds_after_set_project",
                                                            is_required=False, default_value=False)):
        if 'dds_plugin.development_studio' in sys.modules:
            from dds_plugin.development_studio import DevelopmentStudio
            while (True):
                answ = input("Запустить DDS? (y,n):") if confirm else 'y'
                if answ=='y' or answ=='Y':
                    DevelopmentStudio(config_path).run()
                    break
                elif answ=='n' or answ=='N':
                    break
        else:
            log.warning(f'Компонента Directum Development Studio не установлена.')


def repo_info(root_src, folder):
    path = str(PurePath(root_src, folder))

    if pathlib.Path(path).exists():
        stdout_messages: List[str] = ['']
        result = git_tools.git_run("branch --show-current", cwd=path, silent=True, log_stdout=False,
                        filter=process.save_stdout_message_handler(stdout_messages))
        if result == 0:
            branch = stdout_messages.pop()
            if branch == "":
                commit_hash = ""
                if git_tools.git_run("rev-parse HEAD", cwd=path, silent=True, log_stdout=False,
                                     filter=process.save_stdout_message_handler(stdout_messages)) == 0:
                    commit_hash = stdout_messages.pop()
                tag = ""
                stdout_messages_f: List[str] = []
                if git_tools.git_run("show-ref --tags", cwd=path, silent=True, log_stdout=False,
                                     filter=process.save_stdout_message_handler(stdout_messages_f)) == 0:
                    for tag_line in list(filter(lambda x: x.startswith(commit_hash), stdout_messages_f)):
                        tag = tag_line.split(" ")[1][5:] if tag == "" else f'{tag}, {tag_line.split(" ")[1][5:]}'
                if tag == "":
                    detail = f"{commit_hash[:8]}..."
                else:
                    detail = tag
            else:
                detail = branch

            stdout_messages_f: List[str] = []
            result = git_tools.git_run("status -s",
                            cwd=path,
                            filter=process.save_stdout_message_handler(stdout_messages_f),
                            log_stdout=False)
            if result == 0:
                changes_dict = {}
                for m in stdout_messages_f:
                    t = m.split(" ")[0]
                    changes_dict[t] = changes_dict.get(t, 0)+1
                changes = ""
                for k,v in changes_dict.items():
                    if len(changes) == 0:
                        changes = f'{k}:{v}'
                    else:
                        changes = f'{changes}, {k}:{v}'
                return f'({_colorize_green(detail)}) {changes}'
    return f'{_colorize("no data", color="yellow", attrs=["bold"])}'

#region localization helper.
#region work with excel.
def create_for_localization_worksheet(workbook):
    """Создать лист excel для выгрузки ресурсов на локализацию.
    
    Args:
        workbook: книга в excel. 

    Return:
        Лист в excel-файле для выгрузки ресурсов на локализацию. 
    """
    worksheet = workbook.active
    worksheet.title = "На локализацию"
    worksheet.column_dimensions['A'].width = 10
    worksheet['A1'] = "Источник"
    worksheet.column_dimensions['B'].width = 40
    worksheet['B1'] = "Сущность"
    worksheet.column_dimensions['C'].width = 40
    worksheet['C1'] = "Имя ресурса"
    worksheet.column_dimensions['D'].width = 80
    worksheet['D1'] = "Русский текст ресурса"
    worksheet.column_dimensions['E'].width = 80
    worksheet['E1'] = "Английский текст ресурса"
    worksheet.column_dimensions['F'].width = 60
    worksheet['F1'] = "Использование"
    worksheet.column_dimensions['G'].width = 40
    worksheet['G1'] = "Исправленное имя ресурса"
    worksheet.column_dimensions['H'].width = 80
    worksheet['H1'] = "Исправленный русский текст"
    worksheet.column_dimensions['I'].width = 80
    worksheet['I1'] = "Исправленный английский текст"
    worksheet.column_dimensions['J'].width = 80
    worksheet['J1'] = "Примечание"
    worksheet.column_dimensions['K'].width = 80
    worksheet['K1'] = "Вопрос"
    worksheet.column_dimensions['L'].width = 80
    worksheet['L1'] = "Ответ"
    header = worksheet['A1:L1']
    add_style_to_header(header)
    return worksheet

def create_named_worksheet(workbook, worksheet_title):
    """Создать лист excel для выгрузки ресурсов (например, неиспользуемых, с символами из другого языка, с несоответствием пробелов).

    Args:
        workbook: книга в excel.
        worksheet_title: название листа.
    
    Return:
        Лист в excel-файле для выгрузки ресурсов. 
    """
    worksheet = workbook.create_sheet()
    worksheet.title = worksheet_title
    worksheet.column_dimensions['A'].width = 10
    worksheet['A1'] = "Источник"
    worksheet.column_dimensions['B'].width = 40
    worksheet['B1'] = "Сущность"
    worksheet.column_dimensions['C'].width = 40
    worksheet['C1'] = "Имя ресурса"
    worksheet.column_dimensions['D'].width = 80
    worksheet['D1'] = "Русский текст ресурса"
    worksheet.column_dimensions['E'].width = 80
    worksheet['E1'] = "Английский текст ресурса"
    header = worksheet['A1:E1']
    add_style_to_header(header)
    return worksheet

def add_style_to_header(range):
    """Применить форматирование заголовков для заданной области excel-файла.
    
    Args:
        range: область excel-файла.  
    """
    from openpyxl.styles import PatternFill, Alignment, Font
    for row in range:
        for cell in row:
            cell.fill = PatternFill("solid", fgColor="CCCCCC")
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center",
                                       vertical="center")

def add_style_to_range(range):
    """Применить форматирование для заданной области excel-файла.
    
    Args:
        range: область excel-файла.  
    """
    from openpyxl.styles import Alignment
    for row in range:
        for cell in row:
            cell.alignment = Alignment(horizontal='general',
                                       vertical='top',
                                       wrap_text=True)

def get_resources_list_from_xls(input_file, sheet_name, res_count):
    """Загрузить ресурсы решения из xlsx файла в список:

    Args:
        input_file: файл xlsx с вычитанными и переведенными ресурсами для загрузки.
        sheet_name: имя листа xlsx файла с ресурсами.
        res_count: количество строк на листе.

    Return:
        Список словарей с ресурсами из xlsx файла.
    """
    all_resources_list = []
    wb = load_workbook(input_file)
    worksheet = wb[sheet_name]
    for i in range(2, res_count + 2):
        line = {'source': worksheet[f'A{i}'].value,
                'component': worksheet[f'B{i}'].value,
                'code': worksheet[f'C{i}'].value,
                'ru_resource': worksheet[f'D{i}'].value,
                'en_resource': worksheet[f'E{i}'].value,
                'using': worksheet[f'F{i}'].value,
                'new_code': worksheet[f'G{i}'].value,
                'new_ru_resource': worksheet[f'H{i}'].value,
                'new_en_resource': worksheet[f'I{i}'].value}
        all_resources_list.append(line)
    return all_resources_list
#endregion

#region work with files and folders.
def find_all_settings_files(src_folders_list):
    """Получить все json-файлы с настройками из заданной папки. 
    Настройки хранятся в подпапке .Settings в файле *.json.
    
    Args:
        src_folders_list: список папок.
    
    Return:
        Список файлов с настройками бизнесс-процессов.
    """
    import glob
    all_settings_files = []
    for src_folder in src_folders_list:
        settings_files = glob.glob(src_folder + "\\**\\*.Settings\\\ProcessKind\\*.json", recursive=True)
        all_settings_files.extend(settings_files)
    all_settings_files = list(filter(lambda x: x.lower().find("_localization") == -1, all_settings_files))
    return all_settings_files


def find_all_mtd_files(src_folders_list):
    """Получить все mtd-файлы из заданной папки.
    Исключить файлы из папок VersionData, так как для них не будет файлов с ресурсами.
    
    Args:
        src_folders_list: список папок.
    
    Return:
        Список файлов с метаданными ПЧ.
    """
    import glob
    all_mtd_files = []
    for src_folder in src_folders_list:
        mtd_files = glob.glob(src_folder + "\\**\\*.mtd", recursive=True)
        all_mtd_files.extend(mtd_files)
    all_mtd_files = list(filter(lambda x: x.lower().find("versiondata") == -1, all_mtd_files))
    return all_mtd_files

def find_all_ess_resources_files(src_ess_folders_list):
    """Получить все xml-файлы с ресурсами ЛК из заданного списка папок.
    Взять только те, где есть раздел localizedStringValues, так как только в них содержатся ресурсы.
    
    Args:
        src_ess_folders_list: список папок.
    
    Return:
        Список файлов с ресурсами ЛК.
    """
    import glob
    import xmltodict
    resources_files = []
    for src_ess_folder in src_ess_folders_list:
        all_xml_files = glob.glob(src_ess_folder + "\\**\\*.xml", recursive=True)
        for filename in all_xml_files:
            import codecs
            with codecs.open(filename, "r", "utf_8_sig") as f:
                text = f.read()
                try:
                    xml_data = xmltodict.parse(text)
                    xml_data['localizedStringValues']
                except:
                    pass
                else:
                    resources_files.append(filename)
    return resources_files

def find_all_src_files(src_folders_list):
    """Получить все файлы с исходниками из заданной папки.
    Взять *.cs (исключив *.g.cs) и файлы разметки отчетов *.frx
    
    Args:
        src_folders_list: список папок.
    
    Return:
        Список файлов с исходниками ПЧ.  
    """
    import glob
    src_files = []
    for src_folder in src_folders_list:
        cs_files = glob.glob(src_folder + "\\**\\*.cs", recursive=True)
        cs_files = list(filter(lambda x: not x.endswith(".g.cs"), cs_files))
        src_files.extend(cs_files)
        frx_files = glob.glob(src_folder + "\\**\\*.frx", recursive=True)
        src_files.extend(frx_files)
    return src_files

def find_all_ess_src_files(src_ess_folders_list):
    """Получить все xml-файлы с исходниками ЛК из заданного списка папок.
    
    Args:
        src_ess_folders_list: список папок.
    
    Return:
        Список файлов с исходниками ЛК.  
    """
    import glob
    src_files = []
    for src_ess_folder in src_ess_folders_list:
        xml_files = glob.glob(src_ess_folder + "\\**\\*.xml", recursive=True)
        src_files.extend(xml_files)
    return src_files

def get_filename_without_ext(filename):
    """Получить имя файла без расширения.
    
    Args:
        filename: полное имя файла.
    
    Return:
        Имя файла без расширения.  
    """
    return os.path.splitext(filename)[0]

def get_file_path(filename):
    """Получить путь до файла.
    
    Args:
        filename: полное имя файла.
    
    Return:
        Путь до файла.  
    """
    return os.path.dirname(filename)

def get_filename_without_ext_and_src_folder(filename):
    """Получить имя файла без пути и расширения.
    
    Args:
        filename: полное имя файла.
    
    Return:
        Имя файла без пути и расширения.  
    """
    return os.path.basename(get_filename_without_ext(filename))

def get_resource_filename_by_mtd_filename(filename, is_system, is_russian):
    """Получить имя файла с ресурсами по имени mtd-файла и параметрам. 
    
    Args:
        filename: полное имя файла.
        is_system: True - системные, иначе несистемные.
        is_russian: True - русские, иначе английские.
        
    Return:
        Имя файла .resx.
    """
    resource_filename = get_filename_without_ext(filename)
    if is_system:
        resource_filename += "System"
    if is_russian:
        resource_filename += ".ru"
    resource_filename += ".resx"
    if os.path.exists(resource_filename):
        return resource_filename
    else:
        return None
#endregion

#region work with resources.
def get_resources_list_from_mtd_file(mtd_filename, src_list):
    """Получить все ресурсы по mtd-файлу: системные и несистемные.
    
    Args:
        mtd_filename: имя файла с метаданными ПЧ.
        src_list: список с исходниками ПЧ.
    
    Return:
        Список с ресурсами из указанного файла.  
    """
    resources_list = get_resources_list_from_file(mtd_filename, False, src_list)
    system_resources_list = get_resources_list_from_file(mtd_filename, True, src_list)
    resources_list.extend(system_resources_list)
    return resources_list

def get_resources_list_from_file(mtd_filename, is_system, src_list):
    """Получить все ресурсы по mtd-файлу: системные или несистемные.
    
    Args:
        mtd_filename: файл с метаданными ПЧ.
        is_system: True - системные, иначе несистемные.
        src_list: список с исходниками ПЧ
    
    Return:
        Список с ресурсами из указанного файла.  
    """
    resources_list = []
    en_resources_filename = get_resource_filename_by_mtd_filename(mtd_filename, is_system, False)
    ru_resources_filename = get_resource_filename_by_mtd_filename(mtd_filename, is_system, True)
    if en_resources_filename is not None and ru_resources_filename is not None:
        # Сначала найти английские строки. Здесь идем от английских, т.к. файлы с ресурсами автоформируемые и нет необходимости проверять парность.
        en_resources_list = find_all_resources(en_resources_filename)
        for en_resource in en_resources_list:
            resource_code = en_resource['@name']
            en_resource_value = en_resource['value']
            if en_resource_value is not None:
                # Файл с ресурсом лежит в папке, совпадающей с именем компоненты.
                # Исключение - Module.resx, для него необходимо взять имя папки решения.
                splited_path = mtd_filename.split("\\")
                if get_filename_without_ext_and_src_folder(mtd_filename) == 'Module':
                    component_name = splited_path[len(splited_path) - 3]
                else:
                    component_name = splited_path[len(splited_path) - 2]
                # По имени английской строки получить русскую.
                ru_resource = find_resource_in_file_by_code(ru_resources_filename, resource_code)
                ru_resource_value = ""
                if ru_resource is not None:
                    ru_resource_value = ru_resource['value']
                using = ""
                if not is_system:
                    using = find_resource_in_src_data(src_list, resource_code)
                line = {'source': APP_SOURCE ,
                        'filename': mtd_filename,
                        'component': component_name,
                        'code': resource_code,
                        'ru_resource': ru_resource_value,
                        'en_resource': en_resource_value,
                        'is_system': is_system,
                        'using': using,
                        'remark': ""}
                resources_list.append(line)
    return resources_list

def find_all_resources(filename):
    """Получить ресурсы из указанного файла прикладной.
    
    Args:
        filename: файл ресурсами ПЧ.
    
    Return:
        Список с ресурсами из указанного файла.  
    """
    import xmltodict
    import codecs
    with codecs.open(filename, "r", "utf_8_sig") as f:
        text = f.read()
        xml_data = xmltodict.parse(text)
        try:
            resources_list = xml_data['root']['data']
        except:
            return []
        else:
            # Если строка единственная - вместо списка словарей xmltodict вернет словарь,
            # завернуть в список из одного элемента, иначе упадет на получении данных ресурса.
            if not isinstance(resources_list, list):
                resources_list = [resources_list]
            return resources_list

def find_resource_in_file_by_code(filename, resource_code):
    """Найти ресурс в файле по указанному имени.
    
    Args:
        filename: файл ресурсами ПЧ.
        resource_code: код ресурса.
    
    Return:
        Ресурс.  
    """
    import xmltodict
    import codecs
    with codecs.open(filename, "r", "utf_8_sig") as f:
        text = f.read()
        xml_data = xmltodict.parse(text)
        try:
            resources_list = xml_data['root']['data']
        except:
            return None
        else:
            # Если строка единственная - вместо списка словарей вернется словарь,
            # завернуть в список из одного элемента, иначе упадет на фильтрации.
            if not isinstance(resources_list, list):
                resources_list = [resources_list]
            resource = list(filter(lambda x: x['@name'] == resource_code, resources_list))
            if len(resource) >= 1:
                return resource[0]
    return None

def find_all_ess_resources(filename):
    """Получить ресурсы из указанного файла конфигов ЛК.
    
    Args:
        filename: имя файла с ресурсами ЛК.
    
    Return:
        Список с ресурсами из указанного файла.  
    """
    import xmltodict
    import codecs
    with codecs.open(filename, "r", "utf_8_sig") as f:
        text = f.read()
        xml_data = xmltodict.parse(text)
        try:
            resources_list = xml_data['localizedStringValues']['localizedStringValue']
        except:
            return []
        else:
            return resources_list
        
def get_resources_list_from_settings_file(settings_filename):
    """Получить ресурсы из настройки бизнесс-процессов.
    Args:
        settings_resources_filename: файл с настройками бизнес-процессов.

    Return:
        Список с ресурсами.
    """
    resources_list = []   

    # В файле *_properties_localization.json хранятся ресурсы свойств настроек.
    # В файле *_localization.json хранятся ресурсы из настройки схемы бизнесс процессов.
    settings_resources_filename = settings_filename.replace(".json", "_localization.json")
    settings_properties_resources_filename = settings_filename.replace(".json", "_properties_localization.json")
    import os.path
    if os.path.exists(settings_resources_filename) == False:
        return resources_list
    import codecs 
    try:
        with codecs.open(settings_properties_resources_filename, "r", encoding='utf-8') as manifest_json:
            data = " ".join(manifest_json.readlines())
            manifest_dict = json.loads(data)
            component = manifest_dict["Name"]["ru-RU"] 
    except:
        component = ""

    # Получить названия секций с ресурсами из файла со схемой.
    
    with codecs.open(settings_filename, "r", encoding='utf-8') as manifest_json:
        data = " ".join(manifest_json.readlines())
        manifest_dict = json.loads(data)
        shema = manifest_dict["Scheme"]["Scheme"]

    import xml.etree.ElementTree as ET
    root = ET.fromstring(shema)
    resourse_code_list = root.findall(".//**[Name='localizationStringId']//Value")
   

    with codecs.open(settings_resources_filename, "r", encoding='utf-8-sig') as manifest_json:
        data = " ".join(manifest_json.readlines())
        manifest_dict = json.loads(data)
        for resourse_code in resourse_code_list:
            try:
                line = {'source': SETTINGS_SOURCE,
                       'filename': settings_resources_filename,
                        'component': component,
                        'code': resourse_code.text,
                        'ru_resource': manifest_dict[resourse_code.text]["ru-RU"],
                        'en_resource': manifest_dict[resourse_code.text]["default"],
                        'is_system': False,
                        'using': settings_resources_filename,
                        'remark': ""}
                resources_list.append(line)
            except:
                pass
     
    return resources_list

def get_resources_list_from_ess_xml_file(ess_resources_filename, src_ess_list):
    """Получить ресурсы из конфига ЛК.
    
    Args:
        ess_resources_filename: файл с ресурсами ЛК.
        src_ess_list: список с исходниками ЛК.
    
    Return:
        Список с ресурсами.  
    """
    resources_list = []
    all_resources_list = find_all_ess_resources(ess_resources_filename)
    code_resources_list = list(set(map(lambda x: x['@code'], all_resources_list)))
    for resource_code in code_resources_list:
        find_en_resource = find_ess_resource_in_list(all_resources_list, resource_code, "en")
        find_ru_resource = find_ess_resource_in_list(all_resources_list, resource_code, "ru")
        using = find_resource_in_ess_src_data(src_ess_list, resource_code)
        line = {'source': ESS_SOURCE,
                'filename': ess_resources_filename,
                'component': get_filename_without_ext_and_src_folder(ess_resources_filename),
                'code': resource_code,
                'ru_resource': find_ru_resource['resource'],
                'en_resource': find_en_resource['resource'],
                'is_system': False,
                'using': using,
                'remark': f"{find_en_resource['remark']} {find_ru_resource['remark']}".strip()}
        resources_list.append(line)
    return resources_list

def find_ess_resource_in_list(all_resources_list, resource_code, language):
    """Получить ресурс из списка по коду на указанном языке.
    
    Args:
        all_resources_list: список с ресурсами ЛК.
        resource_code: код ресурса.
        language: язык.
    
    Return:
        Словарь из ресурса на указанном языке и примечания.  
    """
    remark = ''
    resource = ''
    resource_list_by_code = list(filter(lambda x: x['@language'].lower() == language.lower() and
                                                  x['@code'].lower() == resource_code.lower(), all_resources_list))
    finded_resource_count = len(resource_list_by_code)
    if finded_resource_count >= 1:
        resource = get_resource_ess_text(resource_list_by_code[0])
        if finded_resource_count > 1:
            remark = f"Найдено больше одной строки с кодом '{resource_code}' на языке '{language}'."
    else:
        resource = ''
        remark = f"Не найдена строка с кодом '{resource_code}' на языке '{language}'."
    return {'resource': resource, 'remark': remark}

def get_resource_ess_text(resource):
    """Получить текст ресурса из конфига ЛК.
    Если текст не указан - вернетсяя пустая строка.
    
    Args:
        resource: ресурс из конфига ЛК.
    
    Return:
        Текст ресурса.  
    """
    try:
        resource_text = resource['#text']
    except:
        resource_text = ''
    return resource_text

def get_resources_list_from_src(src_folders_list, src_ess_folders_list):
    """Получить полный список всех ресурсов из файлов решения: из ПЧ и из ЛК.
    
    Args:
        src_folders_list: список папок с исходниками ПЧ.
        src_ess_folders_list: список папок с конфигами ЛК.
    Return:
        Список с ресурсами.
    """
    resources_list = []
    # Выгрузить ресурсы ПЧ из исходников.
    mtd_files_list = find_all_mtd_files(src_folders_list)
    src_list = get_all_src_data(src_folders_list)
    for mtd_filename in mtd_files_list:
        log.info(mtd_filename)
        resources_list_in_file = get_resources_list_from_mtd_file(mtd_filename, src_list)
        resources_list.extend(resources_list_in_file)
    # Выгрузить ресурсы ЛК из исходников.
    ess_resources_files_list = find_all_ess_resources_files(src_ess_folders_list)
    src_ess_list = get_all_ess_src_data(src_ess_folders_list)
    for ess_resources_filename in ess_resources_files_list:
        log.info(ess_resources_filename)
        resources_list_in_file = get_resources_list_from_ess_xml_file(ess_resources_filename, src_ess_list)
        resources_list.extend(resources_list_in_file)
    # Выгрузить ресурсы из настроек схем бизнес-процессов.
    settings_resourses_files_list = find_all_settings_files(src_folders_list)
    for settings_resourses_filename in settings_resourses_files_list:
        log.info(settings_resourses_filename)
        resources_list_in_file = get_resources_list_from_settings_file(settings_resourses_filename)
        resources_list.extend(resources_list_in_file)
    return resources_list
#endregion

# region search in sources functions.
def get_all_src_data(src_folders_list):
    """Получить исходники прикладной из указанной папки.
    
    Args:
        src_folders_list: список папок с конфигами ЛК.
    
    Return:
        Список с исходниками ПЧ из указанной папки.
    """
    import codecs
    src_data = []
    src_files = find_all_src_files(src_folders_list)
    for filename in src_files:
        with codecs.open(filename, "r", "utf_8_sig") as f:
            try:
                text = f.read()
            except:
                pass
            else:
                src_data.append({'filename': filename, 'text': text})
    return src_data

def find_resource_in_src_data(src_data, resource_code):
    """Найти использование ресурса в исходниках прикладной из указанной папки.
    
    Args:
        src_data: список с исходниками.
        resource_code: код ресурса.
    
    Return:
        Фрагмент с использованием ресурса.
    """
    import re
    pattern = f'Converter\(\"{resource_code}\"\)|Resources\s*\.\s*{resource_code}|Resources\s*\.\s*\w+Report\s*\.\s*{resource_code}'.lower()
    for data in src_data:
        filename = data['filename']
        text = data['text']
        res = re.search(pattern, text.lower())
        if res is not None:
            # Разбить файл построчно, найти номер строки, в которой есть код ресурса.
            src_file_list = text.split("\n")
            str_number = list(map(lambda x: x.lower().find(resource_code.lower()) != -1, src_file_list)).index(True)
            # Вывести фрагмент использования.
            return get_using_fragment(src_file_list, filename, str_number)
    return ""

def get_all_ess_src_data(src_ess_folders_list):
    """Получить исходники ЛК из указанной папки.
    
    Args:
        src_ess_folders_list: список папок с конфигами ЛК.
    
    Return:
        Список с исходниками ЛК из указанной папки.
    """
    import codecs
    src_data = []
    src_files = find_all_ess_src_files(src_ess_folders_list)
    for filename in src_files:
        with codecs.open(filename, "r", "utf_8_sig") as f:
            try:
                text = f.read()
            except:
                pass
            else:
                src_data.append({'filename': filename, 'text': text})
    return src_data

def find_resource_in_ess_src_data(src_list, resource_code):
    """Найти использование ресурса в исходниках.
    
    Args:
        src_ess_list: список с исходниками.
        resource_code: код ресурса.
    
    Return:
        Место использования. Первое вхождение.
    """
    import re
    pattern = f"\.{resource_code}".lower()
    for data in src_list:
        filename = data['filename']
        text = data['text']
        res = re.search(pattern, text.lower())
        if res is not None:
            # Разбить файл построчно, найти номер строки, в которой есть код ресурса.
            src_file_list = text.split("\n")
            str_number = list(map(lambda x: x.lower().find(f".{resource_code}".lower()) != -1, src_file_list)).index(True)
            # Вывести фрагмент использования.
            return get_using_fragment(src_file_list, filename, str_number)
    return ""

def get_using_fragment(src_file_list, filename, str_number):
    """Найти фрагмент использования ресурса.
    
    Args:
        src_file_list: файл с исходниками, разбитый на список строк.
        filename: путь до файла с исходниками.
        str_number: номер строки.
    
    Return:
        Фрагмент использования: файл с папкой, в которой он расположен (полный путь выводить не имеет смысла), номер строки, а также сама строка, 
        ее предшествующая и следующая без лишних пробелов.
    """
    s_path = filename.split("\\")
    short_path = f"{s_path[len(s_path) - 2]}\{s_path[len(s_path) - 1]}"
    src_file_fragment = ""
    if str_number > 0:
        src_file_fragment += f"\n{src_file_list[str_number - 1].strip()}"
    src_file_fragment += f"\n{src_file_list[str_number].strip()}" 
    if str_number < len(src_file_list) - 1:
        src_file_fragment += f"\n{src_file_list[str_number + 1].strip()}"
    return f"{short_path}: {str_number + 1}\n{src_file_fragment}"

def replace_xml_spec_symbols(resource):
    """Заменить спец.символы "<" и ">" в переданном ресурсе. Вспомогательная ф-я.

    Args:
        resource: текст ресурса.
    
    Return:
        Текст ресурса с замененными спец.символами."""
    return resource.replace("<", "&lt;").replace(">", "&gt;")

def import_mtd_resources(src_folders_list, resources_list):
    """Загрузить ресурсы ПЧ в исходники:

    Args:
        src_folders_list: список папок с исходниками ПЧ.
        resources_list: список вычитанных и переведенных строк. 
    """
    import codecs
    # Получить все файлы с ресурсами ПЧ.
    all_mtd_files = find_all_mtd_files(src_folders_list)
    for mtd_filename in all_mtd_files:
        log.info(mtd_filename)
        # Файл с ресурсом лежит в папке, совпадающей с именем компоненты - получить ее для ограничения списка строк для текущего mtd-файла.
        # Исключение - Module.resx, для него необходимо взять имя папки решения.
        splited_path = mtd_filename.split("\\")
        if get_filename_without_ext_and_src_folder(mtd_filename) == 'Module':
            component = splited_path[len(splited_path) - 3]
        else:
            component = splited_path[len(splited_path) - 2]
        mtd_resources_list = list(filter(lambda x: x['component'] == component, resources_list))
        if len(mtd_resources_list) > 0:
            # Обработать сначала английские ресурсы.
            en_res_file_list = [get_resource_filename_by_mtd_filename(mtd_filename, True, False), get_resource_filename_by_mtd_filename(mtd_filename, False, False)]
            for filename in en_res_file_list:
                if os.path.exists(filename):
                    is_modified = False
                    text = ""
                    with codecs.open(filename, "r", "utf_8_sig") as f:
                        text = f.read()
                    res_file_list = text.split("\n")
                    for resource in mtd_resources_list:
                        for index in range(0, len(res_file_list) - 1):
                            name_line = res_file_list[index]
                            value_line = res_file_list[index + 1]
                            if resource['new_en_resource'] is not None and resource['new_en_resource'] != resource['en_resource'] and name_line.find(f'<data name="{resource["code"]}"') != -1:
                                # Обработать однострочные ресурсы отдельно, многострочные отдельно.
                                is_modified = True
                                if value_line.find('<value>') != -1 and value_line.find('</value>') != -1:
                                    res_file_list[index + 1] = value_line.replace(f'>{resource["en_resource"]}<', f'>{replace_xml_spec_symbols(resource["new_en_resource"])}<')
                                else:
                                    res_file_list[index + 1] = f'    <value>{replace_xml_spec_symbols(resource["new_en_resource"])}</value>'
                                    n = 1
                                    while res_file_list[index + n + 1].find("</value>") == -1:
                                        res_file_list[index + n + 1] = DELETE_MARKER
                                        n += 1
                                    res_file_list[index + n + 1] = DELETE_MARKER
                            # Заменить код ресурса.
                            if resource['new_code'] is not None and resource['new_code'] != resource['code'] and name_line.find(f'<data name="{resource["code"]}"') != -1:
                                is_modified = True
                                res_file_list[index] = name_line.replace(f'"{resource["code"]}"', f'"{resource["new_code"]}"')
                    # Сохранять имеет смысл только измененные файлы.
                    if is_modified:
                        with codecs.open(filename, "w", "utf_8_sig") as f:
                            f.write("\n".join(list(filter(lambda x: x.find(DELETE_MARKER) == -1, res_file_list))))
                            #f.write("\n".join(list(res_file_list)))
            # Обработать русские ресурсы.
            ru_res_file_list = [get_resource_filename_by_mtd_filename(mtd_filename, True, True), get_resource_filename_by_mtd_filename(mtd_filename, False, True)]
            for filename in ru_res_file_list:
                if os.path.exists(filename):
                    is_modified = False
                    text = ""
                    with codecs.open(filename, "r", "utf_8_sig") as f:
                        text = f.read()
                    res_file_list = text.split("\n")
                    for resource in mtd_resources_list:
                        for index in range(0, len(res_file_list) - 1):
                            name_line = res_file_list[index]
                            value_line = res_file_list[index + 1]
                            if resource['new_ru_resource'] is not None and resource['new_ru_resource'] != resource['ru_resource'] and name_line.find(f'<data name="{resource["code"]}"') != -1:
                                # Обработать однострочные ресурсы отдельно, многострочные отдельно.
                                is_modified = True
                                if value_line.find('<value>') != -1 and value_line.find('</value>') != -1:
                                    res_file_list[index + 1] = value_line.replace(f'>{resource["ru_resource"]}<', f'>{replace_xml_spec_symbols(resource["new_ru_resource"])}<')
                                else:
                                    res_file_list[index + 1] = f'    <value>{replace_xml_spec_symbols(resource["new_ru_resource"])}</value>'
                                    n = 1
                                    while res_file_list[index + n + 1].find("</value>") == -1:
                                        res_file_list[index + n + 1] = DELETE_MARKER
                                        n += 1
                                    res_file_list[index + n + 1] = DELETE_MARKER
                            # Заменить код ресурса.
                            if resource['new_code'] is not None and resource['new_code'] != resource['code'] and name_line.find(f'<data name="{resource["code"]}"') != -1:
                                is_modified = True
                                res_file_list[index] = name_line.replace(f'"{resource["code"]}"', f'"{resource["new_code"]}"')
                    # Сохранять имеет смысл только измененные файлы.
                    if is_modified:
                        with codecs.open(filename, "w", "utf_8_sig") as f:
                            f.write("\n".join(list(filter(lambda x: x.find(DELETE_MARKER) == -1, res_file_list))))
                            #f.write("\n".join(list(res_file_list)))

def import_ess_resources(src_ess_folders_list, ess_resources_list):
    """Загрузить ресурсы ЛК в исходники:

    Args:
        src_ess_folders_list: список папок с конфигами ЛК.
        ess_resources_list: список вычитанных и переведенных строк. 
    """
    import codecs
    files_list = find_all_ess_resources_files(src_ess_folders_list)
    for filename in files_list:
        log.info(filename)
        component = get_filename_without_ext_and_src_folder(filename)
        resources_list = list(filter(lambda x: x['component'] == component, ess_resources_list))
        if len(resources_list) > 0:
            is_modified = False
            text = ""
            with codecs.open(filename, "r", "utf_8_sig") as f:
                text = f.read()
            res_file_list = text.split("\n")
            for resource in resources_list:
                for index in range(0, len(res_file_list)):
                    line = res_file_list[index]
                    # Обработать английский ресурс, если его меняли.
                    if resource['new_en_resource'] is not None and resource['new_en_resource'] != resource['en_resource'] and line.find(f'code="{resource["code"]}" language="EN"') != -1:
                        # Обработать однострочные ресурсы отдельно, многострочные отдельно.
                        is_modified = True
                        if line.find('</localizedStringValue>') != -1:
                            res_file_list[index] = line.replace(f'>{resource["en_resource"]}<', f'>{replace_xml_spec_symbols(resource["new_en_resource"])}<')
                        else:
                            res_file_list[index] = line.replace('>', f'>\n    {replace_xml_spec_symbols(resource["new_en_resource"])}')
                            n = 1
                            while res_file_list[index + n].find("</localizedStringValue>") == -1:
                                res_file_list[index + n] = DELETE_MARKER
                                n += 1
                    # Обработать русский ресурс, если его меняли.
                    if resource['new_ru_resource'] is not None and resource['new_ru_resource'] != resource['ru_resource'] and line.find(f'code="{resource["code"]}" language="RU"') != -1:
                        # Обработать однострочные ресурсы отдельно, многострочные отдельно.
                        is_modified = True
                        if line.find('</localizedStringValue>') != -1:
                            res_file_list[index] = line.replace(f'>{resource["ru_resource"]}<', f'>{replace_xml_spec_symbols(resource["new_ru_resource"])}<')
                        else:
                            res_file_list[index] = line.replace('>', f'>\n    {replace_xml_spec_symbols(resource["new_ru_resource"])}')
                            n = 1
                            while res_file_list[index + n].find("</localizedStringValue>") == -1:
                                res_file_list[index + n] = DELETE_MARKER
                                n += 1
                    # Заменить коды ресурсов в строках на обоих языках.
                    #if resource['new_code'] is not None and resource['new_code'] != resource['code'] and line.find(f'code="{resource["code"]}"') != -1:
                    #    is_modified = True
                    #    res_file_list[index] = line.replace(f'"{resource["code"]}"', f'"{resource["new_code"]}"')
            # Сохранять имеет смысл только измененные файлы.
            if is_modified:
                with codecs.open(filename, "w", "utf_8_sig") as f:
                    #f.write("\n".join(list(filter(lambda x: x.find(DELETE_MARKER) == -1, res_file_list))))
                    f.write("\n".join(list(res_file_list)))

def import_settings_resources(src_settings_folders_list, settings_resources_list):
    """Загрузить ресурсы настроек схем бизнес-процессов в исходники:

    Args:
        src_settings_folders_list: список папок с исходниками.
        settings_resources_list: список вычитанных и переведенных строк. 
    """
    import codecs
    files_list = find_all_settings_files(src_settings_folders_list)
    for filename in files_list:
        filename = filename.replace(".json", "_localization.json")
        log.info(filename)
        resources_list = list(filter(lambda x: x['using'] == filename, settings_resources_list))
        if len(resources_list) > 0:
            is_modified = False
            with codecs.open(filename, "r", "utf_8_sig") as manifest_json:
                data = " ".join(manifest_json.readlines())
                res_file_list = json.loads(data)
            for resource in resources_list:
                # Обработать английский ресурс, если его меняли.
                if resource['new_en_resource'] is not None and resource['new_en_resource'] != resource['en_resource']:
                    is_modified = True
                    res_file_list[resource['code']]['default']=resource['new_en_resource']
      
                # Обработать русский ресурс, если его меняли.
                if resource['new_ru_resource'] is not None and resource['new_ru_resource'] != resource['ru_resource']:
                    is_modified = True
                    res_file_list[resource['code']]['ru-RU']=resource['new_ru_resource']

            # Сохранять имеет смысл только измененные файлы.
            if is_modified:
                with codecs.open(filename, "w", "utf_8_sig") as f:
                    json.dump(res_file_list, f, indent=4, ensure_ascii=False)

# endregion

#region hight-level functions.
def export_resources(src_folders_list, src_ess_folders_list, is_todo, output_file):
    """Выгрузить ресурсы решения:

    Args:
        src_folders_list: список папок с исходниками ПЧ.
        src_ess_folders_list: список папок с конфигами ЛК.
        is_todo: True - ресурсы с todo, иначе все ресурсы.
        output_file: файла xlsx для выгрузки.
    """
    log.info("==========Экспорт запущен==========")
    log.info("Анализ")
    all_resources_list = get_resources_list_from_src(src_folders_list, src_ess_folders_list)
    log.info("Запись в файл")
    wb = Workbook()
    for_localization_worksheet = create_for_localization_worksheet(wb)
    # На лист "На локализацию" добавить все ресурсы или только ресурсы с todo, в завиимости от варианта запуска.
    # Иcключить неиспользуемые ресурсы, но оставить системные ресурсы (так как в коде не используется),
    # операции из истории (так как в прикладном коде их нет, они подхватываются платформой),
    # ресурсы из настроек схем бизнесс-процессов, т.к. они заданы на схеме. 
    for_localization_resources_list = list(filter(lambda x: x['using'] != ""
                                                  or x['is_system']
                                                  or str(x['code']).startswith("Enum_Operation")
                                                  or x['source'] == SETTINGS_SOURCE, all_resources_list))
    if is_todo:
        # Ряд ресурсов не определяется как строка, добавлено явное преобразование, иначе падает на функциях для работы со строками.
        # Вместе со ресурсами с "todo" выгрузить ресурсы с примечаниями - в примечании указана проблема с ресурсом.
        for_localization_resources_list = list(filter(lambda x: (str(x['ru_resource']).lower().startswith("todo") or
                                                                 str(x['en_resource']).lower().startswith("todo") or
                                                                 x['remark'] != ""), for_localization_resources_list))
    for resource in for_localization_resources_list:
        for_localization_worksheet.append([resource['source'], resource['component'], resource['code'], resource['ru_resource'], resource['en_resource'],
                                           resource['using'], '', '', '', resource['remark']])
    for_localization_count = len(for_localization_resources_list)
    range = for_localization_worksheet['A2:L' + str(for_localization_count + 1)]
    add_style_to_range(range)
    # На лист "Не используется" добавить все неиспользуемые ресурсы. Создавать лист, только если есть ресурсы для добавления.
    # Из списка неиспользуемых исключить:
    #   системные ресурсы, так как в коде не используется, 
    #   операции из истории, так как в прикладном коде их нет, они подхватываются платформой,
    #   ресурсы из настроек схем бизнесс-процессов, т.к. они заданы на схеме. 
    not_used_resources_list = list(filter(lambda x: x['using'] == "" 
                                          and not x['is_system'] 
                                          and not str(x['code']).startswith("Enum_Operation")
                                          and not x['source'] == SETTINGS_SOURCE, all_resources_list))
    not_used_count = len(not_used_resources_list)
    if not_used_count > 0:
        not_used_worksheet = create_named_worksheet(wb, "Не используется")
        for resource in not_used_resources_list:
            not_used_worksheet.append([resource['source'], resource['component'], resource['code'], resource['ru_resource'], resource['en_resource']])
        range = not_used_worksheet['A2:E' + str(not_used_count + 1)]
        add_style_to_range(range)
    # На лист "С анг. символами в рус." добавить те, где в русском ресурсе есть английские символы. Создавать лист, только если есть ресурсы для добавления.
    # Исключить все невычитанные. Иcключить неиспользуемые ресурсы, но оставить системные ресурсы.
    import re
    en_in_ru_resources_list = list(filter(lambda x: (x['using'] != "" or x['is_system']) and 
                                            not str(x['ru_resource']).lower().startswith("todo") and not str(x['en_resource']).lower().startswith("todo") and
                                            (re.search("[a-z]+", str(x['ru_resource']).lower()) is not None), all_resources_list))
    en_in_ru_count = len(en_in_ru_resources_list)
    if en_in_ru_count > 0:
        en_in_ru_worksheet = create_named_worksheet(wb, "С анг. символами в рус.")
        for resource in en_in_ru_resources_list:
            en_in_ru_worksheet.append([resource['source'], resource['component'], resource['code'], resource['ru_resource'], resource['en_resource']])
        range = en_in_ru_worksheet['A2:E' + str(en_in_ru_count + 1)]
        add_style_to_range(range)
    # На лист "С рус. символами в анг." добавить те, где в английском ресурсе есть русские символы. Создавать лист, только если есть ресурсы для добавления.
    # Исключить все невычитанные. Иcключить неиспользуемые ресурсы, но оставить системные ресурсы.
    ru_in_en_resources_list = list(filter(lambda x: (x['using'] != "" or x['is_system']) and 
                                            not str(x['ru_resource']).lower().startswith("todo") and not str(x['en_resource']).lower().startswith("todo") and
                                            (re.search("[а-я]+", str(x['en_resource']).lower()) is not None), all_resources_list))
    ru_in_en_count = len(ru_in_en_resources_list)
    if ru_in_en_count > 0:
        ru_in_en_worksheet = create_named_worksheet(wb, "С рус. символами в анг.")
        for resource in ru_in_en_resources_list:
            ru_in_en_worksheet.append([resource['source'], resource['component'], resource['code'], resource['ru_resource'], resource['en_resource']])
        range = ru_in_en_worksheet['A2:E' + str(ru_in_en_count + 1)]
        add_style_to_range(range)
    # На лист "Несоответствие пробелов" добавить ресурсы, которые:
    #  - содержат двойные пробелы
    ###################  - начинаются или оканчиваются c разного кол-ва пробелов НЕ РАБОТАЕТ, xmltodict.parse обрезает пробелы с концов, условие закомментировано
    #  - имеют разное кол-во переносов строк.
    # Создавать лист, только если есть ресурсы для добавления. Исключить все невычитанные. Иcключить неиспользуемые ресурсы, но оставить системные ресурсы.
    from itertools import takewhile
    incorrect_spacing_resources_list = list(filter(lambda x: (x['using'] != "" or x['is_system']) and
                                                   not str(x['ru_resource']).lower().startswith("todo") and not str(x['en_resource']).lower().startswith("todo") and
                                                   (str(x['ru_resource']).find("  ") != -1 or str(x['en_resource']).find("  ") != -1 or
                                                    #list(takewhile(lambda s: s == " ", list(x['ru_resource']))) != list(takewhile(lambda s: s == " ", list(x['en_resource']))) or
                                                    #list(takewhile(lambda s: s == " ", list(x['ru_resource'][::-1]))) != list(takewhile(lambda s: s == " ", list(x['en_resource'][::-1]))) or
                                                    len(str(x['ru_resource']).split("\r\n")) != len(str(x['en_resource']).split("\r\n"))), all_resources_list))
    incorrect_spacing_count = len(incorrect_spacing_resources_list)
    if incorrect_spacing_count > 0:
        incorrect_spacing_worksheet = create_named_worksheet(wb, "Несоответствие пробелов")
        for resource in incorrect_spacing_resources_list:
            # При добавлении в ячейку добавить "", иначе обрежет переносы строк и пробелы с концов ресурсов.
            incorrect_spacing_worksheet.append([resource['source'], resource['component'], resource['code'], f'"{resource["ru_resource"]}"', f'"{resource["en_resource"]}"'])
        range = incorrect_spacing_worksheet['A2:E' + str(incorrect_spacing_count + 1)]
        add_style_to_range(range)
    # На лист "Пустая рус." добавить ресурсы, у которых нет русской локализации. Потенциально это те, которые вообще не локализовывали.
    from itertools import takewhile
    empty_ru_resources_list = list(filter(lambda x: (x['using'] != "" or x['is_system']) and
                                          not str(x['en_resource']).lower().startswith("todo") and
                                          str(x['ru_resource']) == "", all_resources_list))
    empty_ru_count = len(empty_ru_resources_list)
    if empty_ru_count > 0:
        empty_ru_worksheet = create_named_worksheet(wb, "Пустая рус.")
        for resource in empty_ru_resources_list:
            empty_ru_worksheet.append([resource['source'], resource['component'], resource['code'], resource["ru_resource"], resource["en_resource"]])
        range = empty_ru_worksheet['A2:E' + str(empty_ru_count + 1)]
        add_style_to_range(range)
    # На лист "Анг. = рус." добавить ресурсы, у которых русская и английская строка совпадают. Потенциально это те, которые вообще не локализовывали.
    from itertools import takewhile
    eng_eq_ru_resources_list = list(filter(lambda x: (x['using'] != "" or x['is_system']) and
                                           not str(x['ru_resource']).lower().startswith("todo") and not str(x['en_resource']).lower().startswith("todo") and
                                           str(x['en_resource']) == str(x['ru_resource']), all_resources_list))
    eng_eq_ru_count = len(empty_ru_resources_list)
    if eng_eq_ru_count > 0:
        eng_eq_ru_worksheet = create_named_worksheet(wb, "Анг. равна рус.")
        for resource in eng_eq_ru_resources_list:
            eng_eq_ru_worksheet.append([resource['source'], resource['component'], resource['code'], resource["ru_resource"], resource["en_resource"]])
        range = eng_eq_ru_worksheet['A2:E' + str(eng_eq_ru_count + 1)]
        add_style_to_range(range)
    # Сохранить в файл, предварительно создав папку, если ее еще не существует.
    output_folder = get_file_path(output_file)
    os.makedirs(output_folder, exist_ok=True)
    wb.save(output_file)
    log.info(f"Выгружено в файл: {output_file}")
    log.info(f"Не локализовано: {str(for_localization_count)}")
    log.info(f"Не используется: {str(not_used_count)}")
    log.info(f"С анг. символами в рус.: {str(en_in_ru_count)}")
    log.info(f"С рус. символами в анг.: {str(ru_in_en_count)}")
    log.info(f"Несоответствие пробелов: {str(incorrect_spacing_count)}")
    log.info(f"Пустая русская: {str(empty_ru_count)}")
    log.info(f"Анг. равна рус.: {str(eng_eq_ru_count)}")
    log.info("==========Экспорт завершен==========")

def import_resources(src_folders_list, src_ess_folders_list, input_file, sheet_name, res_count):
    """Загрузить ресурсы решения:

    Args:
        src_folders_list: список папок с исходниками ПЧ.
        src_ess_folders_list: список папок с конфигами ЛК.
        input_file: файл xlsx с вычитанными и переведенными ресурсами для загрузки.
        sheet_name: имя листа xlsx файла с ресурсами.
        res_count: количество строк на листе. 
    """
    print("==========Импорт запущен==========")
    log.info("Анализ")
    all_resources_list = get_resources_list_from_xls(input_file, sheet_name, res_count)
    # Обработать только те ресурсы, которые изменились в процессе вычитки и перевода.
    all_resources_list = list(filter(lambda x: (x['new_code'] is not None and x['new_code'] != x['code']) or
                                               (x['new_ru_resource'] is not None and x['new_ru_resource'] != x['ru_resource']) or
                                               (x['new_en_resource'] is not None and x['new_en_resource'] != x['en_resource']), all_resources_list))
    log.info("Запись в исходники")
    # Импортировать ресурсы ПЧ, если они есть.
    resources_list = list(filter(lambda x: x['source'] == APP_SOURCE, all_resources_list))
    if len(resources_list):
        import_mtd_resources(src_folders_list, resources_list)
    # Импортировать ресурсы ЛК, если они есть.
    ess_resources_list = list(filter(lambda x: x['source'] == ESS_SOURCE, all_resources_list))
    if len(ess_resources_list):
        import_ess_resources(src_ess_folders_list, ess_resources_list)
    # Импортировать ресурсы настроек, если они есть.
    settings_resources_list = list(filter(lambda x: x['source'] == SETTINGS_SOURCE, all_resources_list))
    if len(settings_resources_list):
        import_settings_resources(src_folders_list, settings_resources_list)

    print("==========Импорт завершен==========")

#endregion
#endregion

#endregion

@component(alias=MANAGE_APPLIED_PROJECTS_ALIAS)
class ManageAppliedProject(BaseComponent):
    """ Компонент Изменение проекта. """

    #region constructor-destructor
    def __init__(self, config_path: Optional[str] = None) -> None:
        """
        Конструктор.

        Args:
            config_path: Путь к конфигу.
        """
        super(self.__class__, self).__init__(config_path)
        if 'platform_plugin.static_controller' in sys.modules:
            from platform_plugin.static_controller import StaticController # 4.5
        else:
            from sungero_deploy.static_controller import StaticController  # 4.2-4.4
        self._static_controller = StaticController(self.config_path)

    def install(self) -> None:
        """
        Установить компоненту.
        """
        log.info(f'"{self.__class__.__name__}" component has been successfully installed.')
        self._print_help_after_action()

    def uninstall(self) -> None:
        """
        Удалить компоненту.
        """
        log.info(f'"{self.__class__.__name__}" component has been successfully uninstalled.')
        self._print_help_after_action()
    #endregion

    #region manage projects

    def update_config(self, template_config_path: str, confirm: bool = True, need_pause: bool = False):
        """ Изменить config.yml используя шаблон

        Args:
            template_config_path - путь к конфигу, из которого будут браться новые значения
            confirm: признак необходимости выводить запрос на создание проекта. По умолчанию - True
            need_pause: признак необходимости в конце сделать паузу и ожидать нажатия клавиши пользователем. По умолчанию - False
        """
        inst_path = Path(self.config_path).parent.parent
        log.info(f'Корневой каталог текущего инстанса: {str(inst_path)}')

        def _update_CommentedMap(template_config: CommentedMap, dst_config: CommentedMap):
            for k,v in template_config.items():
                if type(v) == CommentedMap:
                    if k in dst_config.keys():
                        if dst_config[k] is not None:
                            _update_CommentedMap(v, dst_config[k])
                        else:
                            dst_config[k] = v
                    else:
                        dst_config[k] = v
                        _update_CommentedMap(v, dst_config[k])
                elif type(v) == CommentedSeq:
                    if k in dst_config.keys():
                        dst_config[k] = v
                    else:
                        dst_config[k] = v
                        _update_CommentedMap(v, dst_config[k])
                else:
                    dst_config[k] = v

        def _show_CommentedMap(template_config: CommentedMap, dst_config: CommentedMap, indent: int = 1):
            indent_template = "  "
            mark = ""
            for k,v in template_config.items():
                if type(v) == CommentedMap:
                    dst_config_next_level = None
                    if dst_config is not None and k in dst_config.keys():
                        mark = ""
                        dst_config_next_level = dst_config[k]
                    else:
                        mark = _colorize_green('[+]')
                    log.info(f"{(indent)*indent_template}{mark}{k}:")
                    _show_CommentedMap(v, dst_config_next_level, (indent+1))
                elif type(v) == CommentedSeq:
                    if k.lower() == "repository":
                        maxlen = 0
                        for repo in v:
                            if maxlen < len(repo.get("@folderName")):
                                maxlen = len(repo.get("@folderName"))
                        for repo in v:
                            log.info(f'{(indent)*indent_template}{_colorize_cyan("[*]")}folder: {_colorize_green(repo.get("@folderName").ljust(maxlen)):} solutiontype: {_colorize_green(repo.get("@solutionType"))}  url: {_colorize_green(repo.get("@url"))}')
                    else:
                        for r in v:
                            log.info(f"{(indent)*indent_template}{r}")
                else:
                    if dst_config is not None and k in dst_config.keys():
                        if v == dst_config[k]:
                            mark = "[.]"
                            value = f"'{v}'"
                        else:
                            mark = _colorize_cyan('[*]')
                            value = f"'{dst_config[k]}' -> '{v}'"
                    else:
                        mark = _colorize_green('[+]')
                        value = f"'{v}'"
                    log.info(f"{(indent)*indent_template}{mark}{k}: {value}")

        log.info(f'Чтение исходного config.yml: {self.config_path}')
        dst_config = yaml_tools.load_yaml_from_file(self.config_path)
        log.info(f'Чтение файла с требуемыми параметрами: {template_config_path}')
        template_config = yaml_tools.load_yaml_from_file(_get_check_file_path(template_config_path))

        log.info(f'{_colorize_green("Предлагаемые изменения config.yml")}')
        log.info(f'Легенда изменений:')
        log.info(f'{_colorize_cyan("  [*] - значение будет изменено")}')
        log.info(f'{_colorize_green("  [+] - значение будет добавлено")}')
        log.info(f'  [.] - текущее значение и предлагаемое совпадают')
        log.info(f'config.yml:')
        _show_CommentedMap(template_config, dst_config)
        answ = input("Изменить config.yml? (y,n):") if confirm else 'y'
        if answ=='y' or answ=='Y':
            _update_CommentedMap(template_config, dst_config)
            yaml_tools.yaml_dump_to_file(dst_config, self.config_path)
        if need_pause or need_pause is None:
            pause()

    def create_project(self, project_config_path: str, package_path:str = "",
                       need_import_src:bool = False, confirm: bool = True,
                       rundds: bool = None, need_pause: bool = False) -> None:
        """ Создать новый прикладной проект (эксперементальная фича).
        Будет создана БД, в неё будет принят пакет разработки и стандратные шаблоны.

        Args:
            project_config_path: путь к файлу с описанием проекта
            package_path: путь к пакету разработки, который должен содержать бинарники
            need_import_src: признак необходимости принять исходники из указанного пакета разработки. По умолчанию - False
            confirm: признак необходимости выводить запрос на создание проекта. По умолчанию - True
            rundds: признак необходимости запускать DDS. По умолчанию - None, т.е. будет браться значение, определенное в config.yml
            need_pause: признак необходимости в конце сделать паузу и ожидать нажатия клавиши пользователем. По умолчанию - False
        """
        while (True):
            """Подгрузить необходимые модули.
            Выполняется именно тут, т.к:
            * если делать при загрузке - то модули-зависимости могут не успеть подгрузиться
            * DDS и DirectumRX может не быть не установлены и надо об этом сообщать
            """
            if 'sungero_deploy.tools.rxcmd' in sys.modules:
                from sungero_deploy.tools.rxcmd import RxCmd
            elif 'rx_plugin.rxcmd' in sys.modules:
                from rx_plugin.rxcmd import RxCmd
            else:
                log.error('Не найден модуль rxcmd')
                raise RuntimeError('Не найден модуль rxcmd')

            _show_config(project_config_path)
            answ = input("Создать новый проект? (y,n):") if confirm else 'y'
            if answ=='y' or answ=='Y':
                # остановить сервисы
                log.info(_colorize_green("Остановка сервисов"))
                all = All(self.config)
                all.down()

                # скорректировать etc\config.yml
                log.info(_colorize_green("Корректировка config.yml"))
                dst_config = _update_sungero_config(project_config_path, self.config_path)
                yaml_tools.yaml_dump_to_file(dst_config, self.config_path)
                time.sleep(2)

                # создать БД
                log.info(_colorize_green("Создать БД"))
                exitcode = SungeroDB(get_config_model(self.config_path)).up()
                if exitcode == -1:
                    log.error(f'Ошибка при создании БД')
                    return

                # поднять сервисы
                log.info(_colorize_green("Подъем сервисов"))
                all2 = All(get_config_model(self.config_path))
                all2.config_up()
                all2.up()
                all2.check()

                # принять пакет разработки в БД
                if package_path != "":
                    log.info(_colorize_green("Прием пакета разработки"))
                    if 'platform_plugin.deployment_tool' in sys.modules:
                        from platform_plugin.deployment_tool import DeploymentTool # 4.5
                    else:
                        from sungero_deploy.deployment_tool import DeploymentTool # 4.2-4.4
                    DeploymentTool(self.config_path).deploy(package = package_path, init = True)

                    # импортировать шаблоны
                    log.info(_colorize_green("Перезапуск сервисов"))
                    all2.down()
                    time.sleep(2)
                    all2.up()
                    all2.check()
                    log.info(_colorize_green("Импорт шаблонов"))
                    RxCmd(get_config_model(self.config_path)).import_templates()

                # обновить конфиги DevelopmentStudio и DeploymentToolUI
                # Подгрузка модулей выполняется именно тут, т.к:
                #   * если делать при загрузке - то модули-зависимости могут не успеть подгрузиться
                #   * DevelopmentStudio может не быть не установлены и надо об этом сообщать
                log.info(_colorize_green("Обновление конфига DevelopmentStudio"))
                if 'dds_plugin.development_studio' in sys.modules:
                    from dds_plugin.development_studio import DevelopmentStudio
                    DevelopmentStudio(self.config_path).generate_config_settings()
                    # принять пакет разработки с исходниками
                    if need_import_src:
                        log.info(_colorize_green("Прием пакета разработки"))
                        time.sleep(30) #подождать, когда сервисы загрузятся
                        DevelopmentStudio(self.config_path).run(f'--import-package {package_path}')
                else:
                    log.warning('Модуль development_studio plugin-а dds_plugin для компоненты DevelopmentStudio не найден.')
                log.info(_colorize_green("Обновление конфига DeploymentToolUI"))
                if 'dt_ui_plugin.deployment_tool_ui' in sys.modules:
                    from dt_ui_plugin.deployment_tool_ui import DeploymentToolUI
                    DeploymentToolUI(self.config_path).generate_config_settings()
                else:
                    log.warning('Модуль deployment_tool_ui plugin-а dt_ui_plugin для компоненты DeploymentToolUI не найден.')

                log.info("")
                log.info(_colorize_green("Новые параметры:"))
                self.current()
                if need_pause or need_pause is None:
                    pause()

                # запустить DDS
                _run_dds(self.config_path, rundds, confirm)

                break
            elif answ=='n' or answ=='N':
                break

    def set(self, project_config_path: str = None, confirm: bool = True, rundds: bool = None, need_pause: bool = False) -> None:
        """ Переключиться на указанный прикладной проект

        Args:
            project_config_path: путь к файлу с описанием проекта
            confirm: признак необходимости выводить запрос на создание проекта. По умолчанию - True
            rundds: признак необходимости запускать DDS. По умолчанию - None, т.е. будет браться значение, определенное в config.yml
            need_pause: признак необходимости в конце сделать паузу и ожидать нажатия клавиши пользователем. По умолчанию - False
        """

        if project_config_path is None:
            # если конфиг проекта не передали, то попробовать предложить к выбору файлы из каталога, указанного в переменной project_config_path
            config_yaml = yaml_tools.load_yaml_from_file(self.config_path)
            instance_name = config_yaml["variables"]["instance_name"]
            prj_cfg_path = config_yaml["variables"].get("project_config_path", None)
            if prj_cfg_path is None:
                log.error("Переменная project_config_path отсутствует в config.yml")
                return
            if prj_cfg_path == "":
                log.error("Переменная project_config_path config.yml не имеет значения.")
                return
            if not Path(prj_cfg_path).parent.is_dir():
                log.error(f"В переменная project_config_path указан не существующий каталог {prj_cfg_path}.")
                return
            project_configs_folder = PurePath(prj_cfg_path).parent
            show_all_configs = False
            while (True):
                if show_all_configs:
                    filter = "*.yml"
                else:
                    filter = f"{instance_name}_*.yml"
                configs_list = []
                for child in list(Path(project_configs_folder).glob(filter)):
                    configs_list.append(str(child.name))
                i = 1
                for n in configs_list:
                    log.info(f"{i:2}. {n}")
                    i += 1
                answ = input(f"Введите номер (0 - отмена, 99 - {'Файлы для инстанса' if show_all_configs else 'Все файлы'}):")
                if answ.isdigit():
                    selected_index = int(answ)
                else:
                    selected_index = -1
                if selected_index == 99:
                    show_all_configs = not show_all_configs
                elif selected_index == 0:
                    return
                if selected_index >=1 and selected_index <= len(configs_list):
                    project_config_path = Path(project_configs_folder, configs_list[selected_index-1])
                    break

        while (True):
            _show_config(project_config_path)
            answ = input("Переключиться на указанный проект? (y,n):") if confirm else 'y'
            if answ=='y' or answ=='Y':
                # остановить сервисы
                log.info(_colorize_green("Остановка сервисов"))
                all = All(self.config)
                all.down()

                # скорректировать etc\config.yml
                log.info(_colorize_green("Корректировка config.yml"))
                src_config = yaml_tools.load_yaml_from_file(project_config_path)
                dst_config = yaml_tools.load_yaml_from_file(self.config_path)
                dst_config["services_config"]["DevelopmentStudio"]['REPOSITORIES']["repository"]  = src_config["services_config"]["DevelopmentStudio"]['REPOSITORIES']["repository"].copy()
                dst_config["variables"]["purpose"] = src_config["variables"]["purpose"]
                dst_config["variables"]["database"] = src_config["variables"]["database"]
                dst_config["variables"]["home_path"] = src_config["variables"]["home_path"]
                dst_config["variables"]["home_path_src"]  = src_config["variables"]["home_path_src"]
                # костыль по быстрому, чтобы project_config_path была нужного типа
                dst_config["variables"]["project_config_path"]  = dst_config["variables"]["database"]
                dst_config["variables"]["project_config_path"] = project_config_path
                yaml_tools.yaml_dump_to_file(dst_config, self.config_path)
                time.sleep(2)

                # поднять сервисы
                log.info(_colorize_green("Подъем сервисов"))
                all2 = All(get_config_model(self.config_path))
                all2.config_up()
                all2.up()
                all2.check()

                # обновить конфиги DevelopmentStudio и DeploymentToolUI
                # Подгрузка модулей выполняется именно тут, т.к:
                #   * если делать при загрузке - то модули-зависимости могут не успеть подгрузиться
                #   * DevelopmentStudio может не быть не установлены и надо об этом сообщать
                log.info(_colorize_green("Обновление конфига DevelopmentStudio"))
                if 'dds_plugin.development_studio' in sys.modules:
                    from dds_plugin.development_studio import DevelopmentStudio
                    DevelopmentStudio(self.config_path).generate_config_settings()
                else:
                    log.warning('Модуль development_studio plugin-а dds_plugin для компоненты DevelopmentStudio не найден.')
                log.info(_colorize_green("Обновление конфига DeploymentToolUI"))
                if 'dt_ui_plugin.deployment_tool_ui' in sys.modules:
                    from dt_ui_plugin.deployment_tool_ui import DeploymentToolUI
                    DeploymentToolUI(self.config_path).generate_config_settings()
                else:
                    log.warning('Модуль deployment_tool_ui plugin-а dt_ui_plugin для компоненты DeploymentToolUI не найден.')

                log.info("")
                log.info(_colorize_green("Новые параметры:"))
                self.current()
                if need_pause or need_pause is None:
                    pause()

                # запустить DDS
                _run_dds(self.config_path, rundds, confirm)

                break
            elif answ=='n' or answ=='N':
                break

    def generate_empty_project_config(self, new_config_path: str) -> None:
        """ Создать новый файл с описанием проекта

        Args:
            new_config_path: путь к файлу, который нужно создать
        """
        template_config="""# ключевые параметры проекта
variables:
    # Назначение проекта
    purpose: '<Назначение проекта>'
    # БД проекта
    database: '<База данных>'
    # Домашняя директория, относительно которой хранятся все данные сервисов.
    # Используется только в конфигурационном файле.
    home_path: '<Домашний каталог>'
    # Корневой каталог c репозиториями проекта
    home_path_src: '<корневой каталог репозитория проекта>'
# репозитории
services_config:
    DevelopmentStudio:
        REPOSITORIES:
            repository:
            -   '@folderName': '<папка репозитория-1>'
                '@solutionType': 'Work'
                '@url': '<url репозитория-1>'
            -   '@folderName': '<папка репозитория-2>'
                '@solutionType': 'Base'
                '@url': '<url репозитория-2>'
"""
        _generate_empty_config_by_template(new_config_path, template_config)

    def clone_project(self, src_project_config_path: str, dst_project_config_path: str,
                        confirm: bool = True, rundds: bool = None, need_pause: bool = False) -> None:
        """ Сделать копию прикладного проекта (эксперементальная фича).
        Будет сделана копия БД и домашнего каталога проекта.

        Args:
            src_project_config_path: путь к файлу с описанием проекта-источника
            dst_project_config_path: путь к файлу с описанием проекта, в который надо скопировать
            confirm: признак необходимости выводить запрос на создание проекта. По умолчанию - True
            rundds: признак необходимости запускать DDS. По умолчанию - None, т.е. будет браться значение, определенное в config.yml
            need_pause: признак необходимости в конце сделать паузу и ожидать нажатия клавиши пользователем. По умолчанию - False
        """
        sungero_db = SungeroDB(get_config_model(self.config_path))

        src_project_config = yaml_tools.load_yaml_from_file(_get_check_file_path(src_project_config_path))
        src_sungero_config = _update_sungero_config(src_project_config_path, self.config_path)
        src_dbname = src_project_config["variables"]["database"]
        src_homepath = src_project_config["variables"]["home_path"]
        if not Path(src_homepath).is_dir():
            raise AssertionError(f'Исходный домашний каталог "{src_homepath}" не существует.')
        if not sungero_db.is_db_exist(src_dbname):
            raise AssertionError(f'Исходная база данных "{src_dbname}" не существует.')

        dst_project_config = yaml_tools.load_yaml_from_file(_get_check_file_path(dst_project_config_path))
        dst_dbname = dst_project_config["variables"]["database"]
        dst_homepath = dst_project_config["variables"]["home_path"]
        if Path(dst_homepath).is_dir():
            raise AssertionError(f'Целевой домашний каталог "{dst_homepath}" уже существует.')
        if sungero_db.is_db_exist(dst_dbname):
            raise AssertionError(f'Целевая база данных "{dst_dbname}" уже существует.')

        datadase_engine = src_sungero_config["common_config"]["DATABASE_ENGINE"]
        while (True):
            log.info('')
            log.info(Bold(f'Параметры клонирования проекта:'))
            log.info(f'database: {_colorize_green(src_dbname)} -> {_colorize_green(dst_dbname)}')
            log.info(f'homepath: {_colorize_green(src_homepath)} -> {_colorize_green(dst_homepath)}')

            answ = input("Клонировать проект? (y,n):") if confirm else 'y'
            if answ=='y' or answ=='Y':
                # Копирование БД
                log.info(_colorize_green(f'Копирование базы данных {src_dbname} в {dst_dbname}'))
                if datadase_engine == 'mssql':
                    _copy_database_mssql(self.config, src_dbname, dst_dbname)
                else:
                    _copy_database_postgresql(src_sungero_config, src_dbname, dst_dbname)
                # Сделать копию домашнего каталога проекта
                log.info(_colorize_green(f'Копирование домашнего каталога {src_homepath} {dst_homepath}'))
                shutil.copytree(src_homepath, dst_homepath)
                # переключить проект
                log.info("")
                self.set(dst_project_config_path, confirm, rundds, need_pause)
                break
            elif answ=='n' or answ=='N':
                break

    def dds_wo_deploy(self, project_config_path: str) -> None:
        """ Запустить DDS для просмотра/редактирования исходников проекта без фактического переключения на него.
        При этом блокируется возможность публикации, чтобы не сломать текущий проект.

        Args:
            project_config_path: путь к файлу с описанием проекта, чьи исходники требуется открыть
         """
        if 'dds_plugin.development_studio' in sys.modules:
            # подготовить временные файлы для временных config.yml и _ConfigSettings.xml
            import tempfile
            dst_config_file_descriptor = tempfile.mkstemp(prefix="map_config_", suffix=".yml")
            config_settings_file_descriptor = tempfile.mkstemp(prefix="map_ConfigSettings_", suffix=".xml")
            os.close(dst_config_file_descriptor[0])
            os.close(config_settings_file_descriptor[0])
            dst_config_path = dst_config_file_descriptor[1]
            config_settings_file_name = config_settings_file_descriptor[1]
            log.info(f"Создан файл для временного config.yml: {dst_config_path}")
            log.info(f"Создан файл для временного _ConfigSettings.xml: {config_settings_file_name}")

            # подготовить специальный config.yml с проектом, чьи исходники надо открыть
            src_config = yaml_tools.load_yaml_from_file(project_config_path)
            dst_config = yaml_tools.load_yaml_from_file(self.config_path)
            dst_config["services_config"]["DevelopmentStudio"]['REPOSITORIES']["repository"]  = src_config["services_config"]["DevelopmentStudio"]['REPOSITORIES']["repository"].copy()
            dst_config["variables"]["purpose"] = src_config["variables"]["purpose"]
            dst_config["variables"]["database"] = src_config["variables"]["database"]
            dst_config["variables"]["home_path"] = src_config["variables"]["home_path"]
            dst_config["variables"]["home_path_src"]  = src_config["variables"]["home_path_src"]
            # отключить возможность публикации
            dst_config["services_config"]["DevelopmentStudio"]["LOCAL_WEB_RELATIVE_PATH"] = ""
            dst_config["services_config"]["DevelopmentStudio"]["LOCAL_SERVER_HTTP_PORT"] = ""
            dst_config["services_config"]["DevelopmentStudio"]["SERVICE_RUNNER_CONFIG_PATH"] = ""
            yaml_tools.yaml_dump_to_file(dst_config, dst_config_path)

            # подготовить специальный _ConfigSettings.xml для DDS
            from dds_plugin.development_studio import DevelopmentStudio
            from sungero_deploy.services_config import generate_service_config, get_default_tool_host_values_mapping
            dds = DevelopmentStudio(dst_config_path)
            generate_service_config(config_settings_file_name, get_config_model(dst_config_path), dds.instance_service,
                                get_default_tool_host_values_mapping())

            # запустить dds со специальным _ConfigSettings.xml
            cmd = f'"{dds._get_exe_path()}" --multi-instance --settings {config_settings_file_name}'
            exit_code = process.try_execute(cmd, encoding='cp1251')

            # удалить файлы с временными конфигами
            log.info("Удаление файлов временных конфигов.")
            os.remove(dst_config_path)
            os.remove(config_settings_file_name)

    #endregion

    #region manage distribution
    def build_distributions(self, distributions_config_path: str, destination_folder: str,
                            repo_folder: str, increment_version: bool = True, need_pause: bool = False) -> int:
        """ Построить дистрибутивы проекта

        Args:
            distributions_config_path: путь к yml-файл, в котором описаны дистрибутивы, которые нужно собрать
            destination_folder: папка, в которой будет создага папка с номером версии, внутри которой будут подготовлены дистрибутивы
            repo_folder: путь к локальному репозиторию, дистрибутивы которого надо собрать
            increment_version: признак необходимости увеличить номер версии решения после сборки дистрибутива
            need_pause: признак необходимости в конце сделать паузу и ожидать нажатия клавиши пользователем. По умолчанию - False
        """
        try:
            # Проверить переданные параметры
            if not Path(distributions_config_path).is_file():
                raise FileNotFoundError(f'Не найдет конфиг описания дистрибутивов проекта {distributions_config_path}')
            if not Path(destination_folder).is_dir():
                raise FileNotFoundError(f'Не найдет каталог назначения {destination_folder}')
            if not Path(PurePath(repo_folder)).is_dir():
                raise FileNotFoundError(f'Не найдет каталог назначения {repo_folder}')

            # загрузить конфиг с описанием дистрибутивов
            distr_config = yaml_tools.load_yaml_from_file(distributions_config_path)

            # достать номер номер версии и инициализиовать папку версии в папке назначения
            mtd_for_version = PurePath(repo_folder, distr_config["mtd_for_version"])
            if not Path(mtd_for_version).is_file():
                raise FileNotFoundError(f'Не найдет mtd-файл для получения версии решения {mtd_for_version}')
            mtd = yaml_tools.load_yaml_from_file(mtd_for_version)
            version = mtd["Version"]
            log.info(_colorize_green(f'Номер версии {version}'))
            version_folder = PurePath(destination_folder, version)
            io_tools._create_or_clean_dir(version_folder)

            # readme_string - массив строк для readme.md, в котором будет перечень дистрибутивов
            readme_strings = []
            readme_strings.append(distr_config["project"])
            readme_strings.append(f'Версия: {version}')
            readme_strings.append(f'Варианты дистрибутивов: ')
            for distr in distr_config["distributions"]:
                log.info(_colorize_green(f'Обработка дистрибутива {distr["id"]}'))
                readme_strings.append(f'* {distr["folder_name"]} - {distr["comment"]}')
                readme_strings.append("")

                # проинициализировать папку дистрибутива
                distr_folder =  PurePath(version_folder, distr["folder_name"])
                io_tools._create_or_clean_dir(distr_folder)
                # выгрузить пакеты разработки, при этом номер версии не увеличивать
                for devpack in distr["devpacks"]:
                    devpack_config = _get_full_path(repo_folder, devpack["config"])
                    if Path(devpack_config).is_file():
                        result_devpack = str(PurePath(distr_folder, devpack["result"]))
                        self.export_devpack(devpack_config, result_devpack, increment_version=False)
                    else:
                        log.warning(f'Не найден XML-конфиг {devpack_config}')
                # скопировать уникальные для дистрибутива файлы и каталоги
                if distr["files"] is not None:
                    for f in distr["files"]:
                        if f["src"] != "":
                            src = _get_full_path(repo_folder, f["src"])
                            dst = PurePath(distr_folder, f["dst"])
                            log.info(_colorize_green(f'  Копирование {src} -> {dst}'))
                            if Path(src).is_file():
                                shutil.copy(str(src), str(dst))
                            elif Path(src).is_dir():
                                shutil.copytree(str(src), str(dst))
                            else:
                                log.warning(f'Не найдет источник "{src}", указанный для дистрибутива {distr["id"]}')
                # скопировать каталоги и файлы, которые дублируются для каждого дистрибутива
                if distr_config["to_every_set"] is not None:
                    for f in distr_config["to_every_set"]:
                        if f["src"] != "":
                            src = _get_full_path(repo_folder, f["src"])
                            dst = PurePath(distr_folder, f["dst"])
                            log.info(_colorize_green(f'  Копирование {src} -> {dst}'))
                            if Path(src).is_file():
                                shutil.copy(str(src), str(dst))
                            elif Path(src).is_dir():
                                shutil.copytree(str(src), str(dst))
                            else:
                                log.warning(f'Не найдет источник "{src}", указанный для всех дистрибутивов')
                # создать архивы дистрибутивов
                if distr["zip_name"] != "":
                    zip_name = str(PurePath(version_folder, f'{distr["zip_name"]} v.{version}.zip'))
                    log.info(_colorize_green(f'Создать архив {zip_name}'))
                    io_tools.create_archive(zip_name, distr_folder)

            # сформировать readme.md для версии
            with open(str(PurePath(version_folder, 'readme.md')), "w", encoding='UTF-8') as f:
                f.write("\n".join(readme_strings))

            # увеличить номер версии, сформировав и удалив указанные пакеты разработки
            if increment_version:
                if distr_config["devpacks_for_increment_version"] is not None:
                    log.info(_colorize_green('Увеличить номер версии решения'))
                    for devpack in distr_config["devpacks_for_increment_version"]:
                        devpack_config = _get_full_path(repo_folder, devpack["config"])
                        if Path(devpack_config).is_file():
                            result_devpack = str(PurePath(version_folder, "__temp_devpack_for_inc_ver.dat"))
                            result_devpack_xml = str(PurePath(version_folder, "__temp_devpack_for_inc_ver.xml"))
                            self.export_devpack(devpack_config, result_devpack, increment_version=True)
                            os.remove(result_devpack)
                            os.remove(result_devpack_xml)
                        else:
                            log.warning(f'Не найден XML-конфиг {devpack_config}')
                else:
                    log.warning(f'Не найден параметр devpacks_for_increment_version - увеличение версии решения не будет выполнено')

            if need_pause or need_pause is None:
                pause()
            return 0
        except Exception as error:
            log.error(f'При формировании дистирибутивов возникла ошибка {error.value}')
            if need_pause or need_pause is None:
                pause()
            return 1

    def export_devpack(self, devpack_config_name: str, devpack_file_name: str,
                       increment_version: bool = None, set_version: str = None,
                       need_pause: bool = False) -> None:
        """Экспортировать пакет разработки

        Args:
            devpack_config_name: имя XML-файла с конфигурацией пакета разработки. Задает параметр --configuration
            devpack_file_name: путь к создаваемому файлу с пакетом разработки. Задает параметр --development-package
            increment_version: признак, который определяет нужно увеличивать номер версии модулей и решений или нет.
            set_version: номер версии, который надо установить. Задает параметр --set-version. . Если указано значение None - то не передается при вызове DDS
            need_pause: признак необходимости в конце сделать паузу и ожидать нажатия клавиши пользователем. По умолчанию - False
        """
        inc_ver_param = ""
        if increment_version is not None:
            inc_ver_param = f'--increment-version {increment_version}'
        set_ver_param = ""
        if set_version is not None:
            set_ver_param = f'--set-version {set_version}'

        """Подгрузить модуль DDS.
        Выполняется именно тут, т.к:
        * если делать при загрузке - то модули-зависимости могут не успеть подгрузиться
        * DDS может не быть не установлены и надо об этом сообщать
        """
        import sys
        if 'dds_plugin.development_studio' in sys.modules:
            from dds_plugin.development_studio import DevelopmentStudio
        else:
            log.error('Не найден модуль dds_plugin.development_studio')
            raise RuntimeError('Не найден модуль dds_plugin.development_studio')
        command = f' --configuration {devpack_config_name} --development-package {devpack_file_name} {inc_ver_param} {set_ver_param}'
        DevelopmentStudio(self.config_path).run(command=command)
        if need_pause or need_pause is None:
            pause()

    def generate_empty_distributions_config(self, new_config_path: str) -> None:
        """ Создать новый файл с описанием дистрибутивов проекта

        Args:
            new_config_path: путь к файлу, который нужно создать
        """
        template_config="""# Название проекта
project: ''

# mtd-файл, из которого берется номер текущей версии
mtd_for_version: '....Solution.Shared\Module.mtd'

# XML-конфиги, которые используются для формирования пакета разработки в процессе увеличения версии решения
devpacks_for_increment_version:
-   config: ''

# Файлы и каталоги, которые копируются в каждый дистрибутив
to_every_set:
-   'src': ''
    'dst': ''

# Описание дистрибутивов
distributions:
    # идентификатор дистритутива
-   'id': ''
    # описание сути дистрибутива
    'comment': ''
    # папка дистрибутива, создается внутри папки версии решения
    'folder_name': ''
    # Значимая часть имени zip-архива с дистрибутивом. Если указать пустую строку - архив не создается
    'zip_name': 'Образец '
    # Пакеты разработки, которые нужно поместить в дистрибутив
    'devpacks':
    -   'config': '.xml'
        'result': '.dat'
    # Уникальные файлы, которые нужно поместить в конкретный дистрибутив
    'files':
    -   'src': ''
        'dst': ''
"""
        _generate_empty_config_by_template(new_config_path, template_config)

    #endregion

    #region other
    def export_res(self, export_res_config: str = None, mode: str = 'todo', output_file: str = None) -> None:
        """Выгрузить ресурсы решения:

        Args:
            export_res_config: путь до конфига, содержащего список папок с исходниками ПЧ и ЛК.
            mode: режим работы: 'todo' - ресурсы с todo, 'all' - все ресурсы.
            output_file: файл xlsx для выгрузки.
        """
        if not os.path.exists(export_res_config):
            raise FileNotFoundError(f'Не найден конфиг, содержащий список папок с исходниками ПЧ и ЛК {export_res_config}')
        export_res_config = yaml_tools.load_yaml_from_file(export_res_config)
        src_folders = export_res_config["variables"]["src_folders"]
        src_ess_folders = export_res_config["variables"]["src_ess_folders"]
        src_folders_list = src_folders.split('|')
        for src_folder in src_folders_list:
            if not os.path.exists(src_folder):
                log.error(f"Папка с исходниками ПЧ {src_folder} не существует.")
                raise FileNotFoundError(f"'src_folder' folder not found: '{src_folder}'")
        src_ess_folders_list = src_ess_folders.split('|')
        for src_ess_folder in src_ess_folders_list:
            if not os.path.exists(src_ess_folder):
                log.error(f"Папка с исходниками ЛК {src_ess_folder} не существует.")
                raise FileNotFoundError(f"'src_ess_folder' folder not found: '{src_ess_folder}'")
        if (mode.lower() != 'todo') and (mode.lower() != 'all'):
            log.error("Режим работы указан неверно: 'todo' - ресурсы с todo, 'all' - все ресурсы.")
            raise ValueError(f"'mode' value '{mode}' incorrect: allowed values are 'todo' or 'all'.")
        if os.path.exists(output_file):
            try:
                os.rename(output_file, output_file + '~')
                os.rename(output_file + '~', output_file)
            except IOError:
                log.error(f"Файл для выгрузки '{output_file}' занят.")
                raise IOError(f"output file '{output_file}' access deny.")
        export_resources(src_folders_list, src_ess_folders_list, mode == 'todo', output_file)

    def import_res(self, import_res_config: str = None, input_file: str = None, sheet_name: str = None, res_count: int = 0) -> None:
        """Загрузить ресурсы решения:

        Args:
            import_res_config: путь до конфига, содержащего список папок с исходниками ПЧ и ЛК.
            input_file: файл xlsx с вычитанными и переведенными ресурсами для загрузки.
            sheet_name: имя листа xlsx файла с ресурсами.
            res_count: количество строк на листе.
        """
        if not os.path.exists(import_res_config):
            raise FileNotFoundError(f'Не найден конфиг, содержащий список папок с исходниками ПЧ и ЛК {import_res_config}')
        import_res_config = yaml_tools.load_yaml_from_file(import_res_config)
        src_folders = import_res_config["variables"]["src_folders"]
        src_ess_folders = import_res_config["variables"]["src_ess_folders"]
        src_folders_list = src_folders.split('|')
        for src_folder in src_folders_list:
            if not os.path.exists(src_folder):
                log.error(f"Папка с исходниками ПЧ {src_folder} не существует.")
                raise FileNotFoundError(f"'src_folder' folder not found: '{src_folder}'")
        src_ess_folders_list = src_ess_folders.split('|')
        for src_ess_folder in src_ess_folders_list:
            if not os.path.exists(src_ess_folder):
                log.error(f"Папка с исходниками ЛК {src_ess_folder} не существует.")
                raise FileNotFoundError(f"'src_ess_folder' folder not found: '{src_ess_folder}'")
        if not os.path.exists(input_file):
            log.error(f"Файл с вычитанными и переведенными ресурсами {input_file} не существует.")
            raise FileNotFoundError(f"'input_file' file not found: '{input_file}'")
        try:
            wb = load_workbook(input_file)
            worksheet = wb[sheet_name]
        except:
            log.error("Указанныйый лист xlsx файла с ресурсами не существует.")
            raise ValueError(f"'sheet_name' xlsx sheet '{sheet_name}' not found.")
        if (res_count <= 0):
            log.error("Количество строк на листе должно быть положительным числом.")
            raise ValueError(f"'res_count' value '{res_count}' must be positive.")
        import_resources(src_folders_list, src_ess_folders_list, input_file, sheet_name, res_count)

    def clear_log(self, root_logs: str = None, limit_day: int = 3, need_pause: bool = False) -> None:
        """Удалить старые логи. Чистит в root_logs и в подкаталогах.
        Предполагается, что последние символы имени файла лога - YYYY-MM-DD.log

        Args:
            root_logs: корневой каталог репозитория. Если не указан, то будут чиститься логи сервисов текущего instance
            limit_day: за сколько последних дней оставить логи. По умолчанию - 3. Если указать 0 - будут удалены все логи.
            need_pause: признак необходимости в конце сделать паузу и ожидать нажатия клавиши пользователем. По умолчанию - False
        """
        if root_logs is None:
            log_folders = []
            for s in self.config.services_config.values():
                if s.get('LOGS_PATH', None) is not None:
                    log_folders.append(s.get('LOGS_PATH', None))
            log_folders = set(log_folders)
        else:
            log_folders = set([root_logs])
        from datetime import datetime, timedelta
        limit_date = (datetime.now() - timedelta(days=limit_day)).strftime("%Y-%m-%d")
        for root_log in log_folders:
            for root, dirs, files in os.walk(root_log):
                for file in files:
                    date_subs = file[-14:-4]
                    if date_subs <= limit_date:
                        os.remove(os.path.join(root, file))
        if need_pause or need_pause is None:
            pause()

    def current(self, need_pause: bool = False) -> None:
        """ Показать параметры текущего проекта

        Args:
            need_pause: признак необходимости в конце сделать паузу и ожидать нажатия клавиши пользователем. По умолчанию - False
        """
        log.info(f'Веб-клиент:          {_get_url(self.config)}')
        _show_config(self.config_path)
        if need_pause or need_pause is None:
            pause()

    def rx_version(self, need_pause: bool = False) -> None:
        """Показать версию RX

        Args:
            need_pause: признак необходимости в конце сделать паузу и ожидать нажатия клавиши пользователем. По умолчанию - False
        """
        ver = _get_rx_version()
        log.info(f'Directum RX: {ver}')
        if need_pause or need_pause is None:
            pause()

    def url(self, need_pause: bool = False) -> None:
        """Показать url для открытия веб-клиента текущего инстанса

        Args:
            need_pause: признак необходимости в конце сделать паузу и ожидать нажатия клавиши пользователем. По умолчанию - False
        """
        log.info(_get_url(self.config))
        if need_pause or need_pause is None:
            pause()

    def check_config(self, config_path: str, need_pause: bool = False) -> None:
        """ Показать содержимое указанного файла описания проекта

        Args:
            config_path: путь к файлу с описанием проекта
            need_pause: признак необходимости в конце сделать паузу и ожидать нажатия клавиши пользователем. По умолчанию - False
        """
        _show_config(config_path)
        if need_pause or need_pause is None:
            pause()

    def check_sdk(self, need_pause: bool = False) -> None:
        """ Проверить наличие необходимых компонент git и .Net

        Args:
            need_pause: признак необходимости в конце сделать паузу и ожидать нажатия клавиши пользователем. По умолчанию - False
        """
        from common_plugin import git_tools
        from py_common import common_paths
        if git_tools.git_run('--version', cwd=common_paths.root_path, log_stdout=False) != 0:
            log.info(f'Git:           {_colorize_red("Not found")}')
        else:
            log.info(f'Git:           {_colorize_green("Ok")}')

        from common_plugin.dotnet_tools import check_path, check_exe, check_dotnet_requirement_version
        result_message = check_path()
        if result_message:
            log.info(f'Path to .Net:  {_colorize_red(result_message)}')
            return
        else:
            log.info(f'Path to .Net:  {_colorize_green("Ok")}')

        result_message = check_exe()
        if result_message:
            log.info(f'dotnet.exe:    {_colorize_red(result_message)}')
            return
        else:
            log.info(f'dotnet.exe:    {_colorize_green("Ok")}')

        result_message = check_dotnet_requirement_version('sdk')
        if result_message:
            log.info(f'Required .Net: {_colorize_red(result_message)}')
        else:
            log.info(f'Required .Net: {_colorize_green("Ok")}')
        if need_pause or need_pause is None:
            pause()

    @staticmethod
    def help() -> None:
        log.info('do map set - переключиться на проект, описаный в указанном yml-файла')
        log.info('do map generate_empty_project_config - создать заготовку для файла описания проекта')
        log.info('do map create_project - создать новый проект: новую БД, хранилище документов, принять пакет разработки, \
инициализировать его и принять стандартные шаблоны')
        log.info('do map update_config - изменить параметры в config.yml взяв значения из переданного файла')
        log.info('do map clone_project - клонировать проект (сделать копии БД и домашнего каталога)')
        log.info('do map dds_wo_deploy - запустить DevelopmentStudio для просмотра/редактирования исходников указанного проекта без возможности публикации')

        log.info('do map build_distributions - сформировать дистрибутивы решения')
        log.info('do map export_devpack - выгрузить пакет разработки')
        log.info('do map generate_empty_distributions_config - сформировать пустой конфиг с описанием дистрибутивов решения')
        log.info('do map clear_log - удалить старые логи')
        log.info('do map current - показать ключевую информацию из текущего config.yml')
        log.info('do map rx_version - показать версию Sungero')
        log.info('do map url - показать url для подключения к веб-клиенту текущего инстанса')
        log.info('do map check_config - показать ключевую информацию из указанного yml-файла описания проекта')
        log.info('do map check_sdk - проверить наличие необходимых компонент git и .Net')

        log.info('do map export_res - выгрузить ресурсы')
        log.info('do map import_res - загрузить ресурсы')

    #endregion
